from __future__ import annotations

import argparse
from itertools import combinations
from math import factorial
from numbers import Number
import re
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCHOOL_STAGE_AGE_RANGES = (
    ("SD", 7, 12),
    ("SMP", 13, 15),
    ("SMA/SMK", 16, 18),
    ("Perguruan Tinggi", 19, 25),
)
SCHOOL_AGE_MIN = SCHOOL_STAGE_AGE_RANGES[0][1]
SCHOOL_AGE_MAX = SCHOOL_STAGE_AGE_RANGES[-1][2]
MISSING_THRESHOLD = 0.20
PHONE_WINSOR_QUANTILE = 0.99

DIMENSION_WEIGHTS = {
    "akses_perangkat": 0.25,
    "konektivitas": 0.25,
    "kapasitas_manusia": 0.20,
    "penggunaan_digital": 0.20,
    "lingkungan_sosial": 0.10,
}

RECOMMENDED_INDICATOR_WEIGHTS = {
    "dimensi_akses_perangkat": {
        "indikator_hp_dimiliki": 1 / 3,
        "indikator_kecukupan_hp": 1 / 3,
        "indikator_perangkat_produktif": 1 / 3,
    },
    "dimensi_konektivitas": {
        "indikator_akses_internet": 1.0,
    },
    "dimensi_kapasitas_manusia": {
        "indikator_pendidikan_kepala": 0.50,
        "indikator_rasio_sekolah": 0.50,
    },
    "dimensi_penggunaan_digital": {
        "indikator_medsos": 1 / 3,
        "indikator_media_informasi": 1 / 3,
        "indikator_partisipasi_kebijakan": 1 / 3,
    },
    "dimensi_lingkungan_sosial": {
        "indikator_organisasi_kepala": 0.25,
        "indikator_organisasi_anggota": 0.25,
        "indikator_partisipasi_masyarakat_kepala": 0.25,
        "indikator_partisipasi_masyarakat_anggota": 0.25,
    },
}

INDICATOR_OUTPUT_MAP = {
    "indikator_hp_dimiliki": "indikator_A",
    "indikator_kecukupan_hp": "indikator_B",
    "indikator_perangkat_produktif": "indikator_C",
    "indikator_akses_internet": "indikator_D",
    "indikator_pendidikan_kepala": "indikator_E",
    "indikator_rasio_sekolah": "indikator_F",
    "indikator_organisasi_kepala": "indikator_G",
    "indikator_organisasi_anggota": "indikator_H",
    "indikator_partisipasi_masyarakat_kepala": "indikator_I",
    "indikator_partisipasi_masyarakat_anggota": "indikator_J",
    "indikator_medsos": "indikator_K",
    "indikator_media_informasi": "indikator_L",
    "indikator_partisipasi_kebijakan": "indikator_M",
}

DIMENSION_OUTPUT_MAP = {
    "dimensi_akses_perangkat": "dimensi_A",
    "dimensi_konektivitas": "dimensi_B",
    "dimensi_kapasitas_manusia": "dimensi_C",
    "dimensi_penggunaan_digital": "dimensi_D",
    "dimensi_lingkungan_sosial": "dimensi_E",
}

ADVANCED_ANALYSIS_DIMENSION_SPECS = (
    {
        "dimension_column": "dimensi_A",
        "dimension_code": "A",
        "weight": DIMENSION_WEIGHTS["akses_perangkat"],
        "indicator_specs": (
            ("indikator_A", "A1"),
            ("indikator_B", "A2"),
            ("indikator_C", "A3"),
        ),
    },
    {
        "dimension_column": "dimensi_B",
        "dimension_code": "B",
        "weight": DIMENSION_WEIGHTS["konektivitas"],
        "indicator_specs": (("indikator_D", "B1"),),
    },
    {
        "dimension_column": "dimensi_C",
        "dimension_code": "C",
        "weight": DIMENSION_WEIGHTS["kapasitas_manusia"],
        "indicator_specs": (
            ("indikator_E", "C1"),
            ("indikator_F", "C2"),
        ),
    },
    {
        "dimension_column": "dimensi_D",
        "dimension_code": "D",
        "weight": DIMENSION_WEIGHTS["penggunaan_digital"],
        "indicator_specs": (
            ("indikator_K", "D1"),
            ("indikator_L", "D2"),
            ("indikator_M", "D3"),
        ),
    },
    {
        "dimension_column": "dimensi_E",
        "dimension_code": "E",
        "weight": DIMENSION_WEIGHTS["lingkungan_sosial"],
        "indicator_specs": (
            ("indikator_G", "E1"),
            ("indikator_H", "E2"),
            ("indikator_I", "E3"),
            ("indikator_J", "E4"),
        ),
    },
)

LOG_ANALYSIS_SCALE = 100.0
LOG_ANALYSIS_EPSILON = 1e-9
OAT_DIMENSION_INCREMENT = 0.01

IID_RT_CATEGORY_RANGES = {
    "sangat rendah": "0,00-0,20",
    "rendah": ">0,20-0,40",
    "sedang": ">0,40-0,60",
    "tinggi": ">0,60-0,80",
    "sangat tinggi": ">0,80-1,00",
}

IID_RT_CATEGORY_ORDER = list(IID_RT_CATEGORY_RANGES.keys())
IID_RT_FIXED_CUTOFFS = (0.20, 0.40, 0.60, 0.80)
UNSCORED_IID_CATEGORY_LABEL = "tanpa skor IID-RT"

GINI_CATEGORY_ORDER = ["Rendah", "Sedang", "Tinggi"]
GINI_INTERPRETATION_RULE_TEXT = (
    "klasifikasi relatif berbasis tertil atas sebaran Gini antar desa dalam sampel; "
    "33,3% desa dengan Gini terendah dikategorikan Rendah, 33,3% berikutnya Sedang, "
    "dan 33,3% dengan Gini tertinggi dikategorikan Tinggi; tidak memakai ambang absolut"
)

EXCEL_FLOAT_FORMAT = "0.######"
EXCEL_PERCENT_FORMAT = "0.######%"

COLUMN_ALIASES = {
    "family_id": ["family_id", "kk", "no_kk", "nokk", "nomor_kartu_keluarga"],
    "abs_id": ["abs_id", "abs id"],
    "subjek": ["subjek", "subyek", "subject"],
    "status_dalam_keluarga": ["status_dalam_keluarga", "status dalam keluarga", "status_keluarga"],
    "kode_bangunan": ["kode_bangunan", "kode bangunan"],
    "kode_deskel": ["kode_deskel", "kode_desa", "kode deskel", "kode desa"],
    "deskel": ["deskel", "desa", "kelurahan", "nama deskel", "nama desa"],
    "dusun": ["dusun"],
    "rw": ["rw"],
    "lat": ["lat", "latitude"],
    "lng": ["lng", "lon", "longitude"],
    "nama": ["nama"],
    "usia": ["usia", "umur"],
    "suku": ["suku"],
    "jml_keluarga": ["jml_keluarga", "jml keluarga", "jumlah keluarga", "jumlah_keluarga"],
    "hp_punya": ["hp_punya", "kepemilikan hp", "kepemilikan_hp", "hp punya"],
    "hp_jumlah": ["hp_jumlah", "jumlah hp", "jumlah_hp"],
    "elektronik_rumah": ["elektronik_rumah", "elektronik rumah", "elektronik"],
    "wifi": ["wifi"],
    "hp_provider": ["hp_provider", "provider hp", "provider_hp", "provider seluler"],
    "rp_komunikasi": ["rp_komunikasi", "rp komunikasi", "biaya komunikasi", "pengeluaran komunikasi"],
    "ijazah": ["ijazah", "pendidikan terakhir", "pendidikan_terakhir", "pendidikan"],
    "partisipasi_sekolah": ["partisipasi_sekolah", "partisipasi sekolah", "status sekolah"],
    "par_organisasi": ["par_organisasi", "partisipasi organisasi", "partisipasi_organisasi", "par_organisa"],
    "organisasi_nama": ["organisasi_nama", "organisasi nama", "nama organisasi"],
    "par_masyarakat": ["par_masyarakat", "partisipasi masyarakat", "partisipasi_masyarakat"],
    "medsos": ["medsos", "media sosial", "media_sosial"],
    "media_informasi": ["media_informasi", "media informasi", "sumber informasi"],
    "par_kebijakan": ["par_kebijakan", "partisipasi kebijakan", "partisipasi_kebijakan"],
}

YES_VALUES = {"ya", "yes", "y", "1", "1.0", "true"}
NO_VALUES = {"tidak", "no", "n", "0", "0.0", "false", "tidak ada"}
PUBLIC_WIFI_VALUES = {
    "tidak berbayar (layanan publik)",
    "tidak berbayar (layanan public)",
    "layanan publik",
    "layanan public",
    "gratis",
    "menumpang",
}
DIGITAL_PRODUCTIVE_DEVICE_KEYWORDS = ("laptop", "notebook", "komputer", "pc")
DIGITAL_MEDIA_KEYWORDS = {"internet", "online", "media sosial", "website", "portal", "aplikasi"}

EDUCATION_SCORE_MAP = {
    "tidak punya ijazah": 0.00,
    "tk/paud": 0.10,
    "tk": 0.10,
    "paud": 0.10,
    "sd/sederajat": 0.25,
    "sd": 0.25,
    "smp/sederajat": 0.50,
    "smp": 0.50,
    "sma/sederajat": 0.75,
    "sma": 0.75,
    "smk/sederajat": 0.75,
    "smk": 0.75,
    "d1/d2/d3": 0.85,
    "d1": 0.85,
    "d2": 0.85,
    "d3": 0.85,
    "d4/s1": 0.95,
    "d4": 0.95,
    "s1": 0.95,
    "s2/s3": 1.00,
    "s2": 1.00,
    "s3": 1.00,
}

SCHOOL_PARTICIPATION_SCORE_MAP = {
    "tidak punya ijazah": 0.00,
    "tk/paud": 0.00,
    "tk": 0.00,
    "paud": 0.00,
    "sd/sederajat": 0.50,
    "sd": 0.50,
    "smp/sederajat": 0.50,
    "smp": 0.50,
    "sma/sederajat": 0.50,
    "sma": 0.50,
    "smk/sederajat": 0.50,
    "smk": 0.50,
    "d1/d2/d3": 1.00,
    "d1": 1.00,
    "d2": 1.00,
    "d3": 1.00,
    "d4/s1": 1.00,
    "d4": 1.00,
    "s1": 1.00,
    "s2/s3": 1.00,
    "s2": 1.00,
    "s3": 1.00,
}


def school_age_mask(
    age_series: pd.Series,
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> pd.Series:
    return age_series.between(school_age_min, school_age_max, inclusive="both")


def is_default_school_age_range(
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> bool:
    return school_age_min == SCHOOL_AGE_MIN and school_age_max == SCHOOL_AGE_MAX


def build_school_stage_summary() -> str:
    return ", ".join(f"{label} {start}-{end} tahun" for label, start, end in SCHOOL_STAGE_AGE_RANGES)


def build_school_age_range_text(
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> str:
    return f"usia {school_age_min}-{school_age_max} tahun"


def build_school_ratio_rule_text(
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> str:
    rule = f"anggota usia {school_age_min}-{school_age_max} yang statusnya sedang sekolah"
    if is_default_school_age_range(school_age_min, school_age_max):
        return f"{rule}; cakupan: {build_school_stage_summary()}"
    return rule


def build_school_indicator_row_text(
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> str:
    return f"anggota {build_school_age_range_text(school_age_min, school_age_max)} saja"


def normalize_column_name(name: str) -> str:
    text = str(name).strip().lower().replace("\u00a0", " ")
    text = re.sub(r"[\s/.-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def normalize_text_series(series: pd.Series) -> pd.Series:
    return (
        series.astype("string")
        .fillna("")
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", " ", regex=True)
    )


def canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed = df.copy()
    renamed.columns = [normalize_column_name(col) for col in renamed.columns]
    rename_map: dict[str, str] = {}
    for canonical_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            normalized_alias = normalize_column_name(alias)
            if normalized_alias in renamed.columns:
                rename_map[normalized_alias] = canonical_name
                break
    renamed = renamed.rename(columns=rename_map)
    for canonical_name in COLUMN_ALIASES:
        if canonical_name not in renamed.columns:
            renamed[canonical_name] = pd.NA
    return renamed


def load_source_data(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, low_memory=False)
    elif path.suffix.lower() == ".parquet":
        df = pd.read_parquet(path)
    elif path.suffix.lower() in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
    else:
        raise ValueError(f"Format file tidak didukung: {path.suffix}")
    return canonicalize_columns(df)


def derive_family_id(df: pd.DataFrame) -> pd.Series:
    family_col = df["family_id"].astype("string").str.strip()
    has_direct_family_id = family_col.notna() & family_col.ne("") & family_col.str.lower().ne("nan")
    if has_direct_family_id.all():
        return family_col

    abs_id = df["abs_id"].astype("string").str.strip()
    parts = abs_id.str.split(".", n=3, expand=True)
    if parts.shape[1] == 4:
        last_segment = parts[3].fillna("")
        household_segment = last_segment.str[:-2].where(last_segment.str.len() > 2, last_segment)
        family_id = (
            parts[0].fillna("")
            + "."
            + parts[1].fillna("")
            + "."
            + parts[2].fillna("")
            + "."
            + household_segment.fillna("")
        )
        family_id = family_id.str.strip(".")
        family_id = family_id.mask(family_id.eq(""), pd.NA)
        if has_direct_family_id.any():
            family_id = family_col.where(has_direct_family_id, family_id)
        return family_id

    generated = pd.Series([f"AUTO_FID_{i + 1:06d}" for i in range(len(df))], index=df.index, dtype="string")
    return generated


def standardize_dusun_label(series: pd.Series) -> pd.Series:
    dusun = normalize_text_series(series)
    dusun = dusun.str.replace(r"^(dusun|dsn)\s+", "", regex=True)
    numeric_only = dusun.str.fullmatch(r"0*\d+")
    dusun_numeric = dusun.where(~numeric_only, dusun.str.lstrip("0"))
    dusun = dusun_numeric.mask(numeric_only & dusun_numeric.eq(""), "0")
    dusun = dusun.replace({"": pd.NA, "nan": pd.NA})
    return dusun


def split_multivalue_text(value: object) -> list[str]:
    if pd.isna(value):
        return []
    items = [item.strip().lower() for item in str(value).split(",")]
    return sorted({item for item in items if item and item != "nan"})


def count_multivalue_items(value: object) -> int:
    return len(split_multivalue_text(value))


def count_keyword_matches(value: object, keywords: tuple[str, ...]) -> int:
    items = split_multivalue_text(value)
    return sum(1 for item in items if any(keyword in item for keyword in keywords))


def collect_multivalue_items(values: pd.Series) -> tuple[str, ...]:
    items: set[str] = set()
    for value in values:
        items.update(split_multivalue_text(value))
    return tuple(sorted(items))


def count_series_multivalue_items(values: pd.Series) -> int:
    return len(collect_multivalue_items(values))


def count_combined_multivalue_items(primary_value: object, secondary_value: object) -> int:
    primary_items = set(split_multivalue_text(primary_value))
    secondary_items = set(split_multivalue_text(secondary_value))
    return max(len(primary_items), len(secondary_items))


def build_head_mask(df: pd.DataFrame) -> pd.Series:
    subjek_norm = normalize_text_series(df["subjek"])
    status_norm = normalize_text_series(df["status_dalam_keluarga"])
    return subjek_norm.eq("kepala keluarga") | status_norm.eq("kepala keluarga")


def score_zero_one_many(value: object) -> float | np.nan:
    if pd.isna(value):
        return np.nan
    numeric_value = float(value)
    if numeric_value <= 0:
        return 0.0
    if numeric_value > 1:
        return 1.0
    return 0.5


def score_binary_presence(value: object) -> float | np.nan:
    if pd.isna(value):
        return np.nan
    return 1.0 if float(value) > 0 else 0.0


def score_capped_ratio(numerator: object, denominator: object) -> float | np.nan:
    if pd.isna(numerator) or pd.isna(denominator):
        return np.nan
    denominator_value = float(denominator)
    if denominator_value <= 0:
        return np.nan
    return float(np.clip(float(numerator) / denominator_value, 0, 1))


def score_internet_access(wifi_items: object, provider_items: object, communication_spend: object) -> float:
    wifi_values = tuple(wifi_items) if isinstance(wifi_items, (list, tuple, set)) else tuple(split_multivalue_text(wifi_items))
    provider_values = (
        tuple(provider_items)
        if isinstance(provider_items, (list, tuple, set))
        else tuple(split_multivalue_text(provider_items))
    )
    wifi_set = {str(item).strip().lower() for item in wifi_values if str(item).strip()}
    provider_set = {str(item).strip().lower() for item in provider_values if str(item).strip()}
    has_wifi = bool(wifi_set)
    has_private_wifi = has_wifi and any(item not in PUBLIC_WIFI_VALUES for item in wifi_set)
    has_public_only_wifi = has_wifi and not has_private_wifi
    has_provider = bool(provider_set)
    spend_value = pd.to_numeric(pd.Series([communication_spend]), errors="coerce").iloc[0]

    if has_private_wifi:
        return 1.0
    if has_provider:
        return 0.75
    if has_public_only_wifi:
        return 0.5
    if pd.notna(spend_value) and float(spend_value) > 0:
        return 0.5
    return 0.0


def score_social_media_use(value: object) -> float:
    item_count = len(split_multivalue_text(value))
    if item_count == 0:
        return 0.0
    if item_count == 1:
        return 0.5
    return 1.0


def score_media_information(value: object) -> float:
    items = split_multivalue_text(value)
    if not items:
        return 0.0
    if any(item in DIGITAL_MEDIA_KEYWORDS or "internet" in item or "online" in item for item in items):
        return 1.0
    return 0.5


def score_education(value: object) -> float | np.nan:
    normalized = normalize_text_series(pd.Series([value])).iloc[0]
    if normalized == "":
        return np.nan
    return EDUCATION_SCORE_MAP.get(normalized, np.nan)


def score_school_participation(value: object) -> float | np.nan:
    normalized = normalize_text_series(pd.Series([value])).iloc[0]
    if normalized == "":
        return np.nan
    return SCHOOL_PARTICIPATION_SCORE_MAP.get(normalized, np.nan)


def format_numeric_label(value: float) -> str:
    return f"{float(value):.6f}".rstrip("0").rstrip(".")


def round_numeric_series(series: pd.Series, digits: int = 6) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return series.round(digits)
    return series


def round_numeric_dataframe(df: pd.DataFrame, digits: int = 6) -> pd.DataFrame:
    rounded = df.copy()
    for column in rounded.columns:
        rounded[column] = round_numeric_series(rounded[column], digits=digits)
    return rounded


def format_output_number(value: object) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, (bool, np.bool_)):
        return str(value)
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        return format_numeric_label(float(value))
    return str(value)


def format_dataframe_for_csv(df: pd.DataFrame) -> pd.DataFrame:
    formatted = df.copy()
    for column in formatted.columns:
        if pd.api.types.is_bool_dtype(formatted[column]):
            continue
        if pd.api.types.is_numeric_dtype(formatted[column]):
            formatted[column] = formatted[column].map(format_output_number)
    return formatted


def classify_iid_rt_with_cutoffs(value: object, cutoffs: tuple[float, float, float, float]) -> object:
    if pd.isna(value):
        return pd.NA
    q1, q2, q3, q4 = cutoffs
    score = float(value)
    if score <= q1:
        return "sangat rendah"
    if score <= q2:
        return "rendah"
    if score <= q3:
        return "sedang"
    if score <= q4:
        return "tinggi"
    return "sangat tinggi"


def build_iid_category_ranges_from_cutoffs(cutoffs: tuple[float, float, float, float]) -> dict[str, str]:
    q1, q2, q3, q4 = cutoffs
    q1_text = format_numeric_label(q1)
    q2_text = format_numeric_label(q2)
    q3_text = format_numeric_label(q3)
    q4_text = format_numeric_label(q4)
    return {
        "sangat rendah": f"<= {q1_text}",
        "rendah": f"> {q1_text}-{q2_text}",
        "sedang": f"> {q2_text}-{q3_text}",
        "tinggi": f"> {q3_text}-{q4_text}",
        "sangat tinggi": f"> {q4_text}",
    }


def compute_weighted_dimension(df: pd.DataFrame, weights: dict[str, float]) -> pd.Series:
    numerator = pd.Series(0.0, index=df.index, dtype=float)
    denominator = pd.Series(0.0, index=df.index, dtype=float)
    for column, weight in weights.items():
        values = pd.to_numeric(df[column], errors="coerce")
        valid_mask = values.notna()
        numerator = numerator + values.fillna(0.0) * weight
        denominator = denominator + valid_mask.astype(float) * weight
    return numerator / denominator.replace(0, np.nan)


def gini_coefficient(values: pd.Series | np.ndarray) -> float:
    array = np.asarray(values, dtype=float)
    array = array[np.isfinite(array)]
    if array.size <= 1:
        return 0.0
    if np.allclose(array, 0):
        return 0.0
    array = np.sort(array)
    n = array.size
    index = np.arange(1, n + 1)
    return float((2 * np.sum(index * array) / (n * np.sum(array))) - ((n + 1) / n))


def classify_iid_rt(value: object) -> object:
    if pd.isna(value):
        return pd.NA
    score = float(value)
    if score <= IID_RT_FIXED_CUTOFFS[0]:
        return "sangat rendah"
    if score <= IID_RT_FIXED_CUTOFFS[1]:
        return "rendah"
    if score <= IID_RT_FIXED_CUTOFFS[2]:
        return "sedang"
    if score <= IID_RT_FIXED_CUTOFFS[3]:
        return "tinggi"
    return "sangat tinggi"


def format_gini_range_value(value: object, digits: int = 5) -> str:
    if pd.isna(value):
        return ""
    return f"{float(value):.{digits}f}"


def apply_relative_gini_classification(
    desa_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    output_columns = [
        "interpretasi_gini",
        "rentang_gini",
        "jumlah_desa",
        "persentase_desa",
        "total_desa",
        "batas_bawah",
        "batas_atas",
    ]
    if desa_summary.empty or "gini_iid_rumah_tangga" not in desa_summary.columns:
        empty_summary = pd.DataFrame(columns=output_columns)
        enriched_df = desa_summary.copy()
        if "interpretasi_gini" not in enriched_df.columns:
            enriched_df["interpretasi_gini"] = pd.NA
        return enriched_df, empty_summary

    enriched_df = desa_summary.copy()
    enriched_df["gini_iid_rumah_tangga"] = pd.to_numeric(enriched_df["gini_iid_rumah_tangga"], errors="coerce")
    enriched_df["interpretasi_gini"] = pd.Series(pd.NA, index=enriched_df.index, dtype="object")

    valid_mask = enriched_df["gini_iid_rumah_tangga"].notna()
    valid_count = int(valid_mask.sum())
    if valid_count > 0:
        tertile_count = min(3, valid_count)
        labels = GINI_CATEGORY_ORDER[:tertile_count]
        ranked_values = enriched_df.loc[valid_mask, "gini_iid_rumah_tangga"].rank(method="first")
        category_values = pd.qcut(ranked_values, q=tertile_count, labels=labels)
        enriched_df.loc[valid_mask, "interpretasi_gini"] = category_values.astype("string").tolist()

    total_desa = int(valid_count)
    category_rows: list[dict[str, object]] = []
    for label in GINI_CATEGORY_ORDER:
        category_values = enriched_df.loc[
            enriched_df["interpretasi_gini"].astype("string").eq(label),
            "gini_iid_rumah_tangga",
        ].dropna()
        jumlah_desa = int(category_values.shape[0])
        batas_bawah = float(category_values.min()) if not category_values.empty else np.nan
        batas_atas = float(category_values.max()) if not category_values.empty else np.nan

        if category_values.empty:
            rentang_gini = pd.NA
        elif label == GINI_CATEGORY_ORDER[0]:
            rentang_gini = f"<= {format_gini_range_value(batas_atas)}"
        elif label == GINI_CATEGORY_ORDER[-1]:
            rentang_gini = f">= {format_gini_range_value(batas_bawah)}"
        else:
            rentang_gini = f"{format_gini_range_value(batas_bawah)} - {format_gini_range_value(batas_atas)}"

        category_rows.append(
            {
                "interpretasi_gini": label,
                "rentang_gini": rentang_gini,
                "jumlah_desa": jumlah_desa,
                "persentase_desa": (jumlah_desa / total_desa) if total_desa > 0 else 0.0,
                "total_desa": total_desa,
                "batas_bawah": batas_bawah,
                "batas_atas": batas_atas,
            }
        )

    return enriched_df, pd.DataFrame(category_rows, columns=output_columns)


def interpret_gini_value(value: object, gini_distribution_df: pd.DataFrame) -> object:
    if pd.isna(value) or gini_distribution_df.empty:
        return pd.NA

    score = float(value)
    lookup_df = gini_distribution_df.copy()
    for column in ("jumlah_desa", "batas_bawah", "batas_atas"):
        if column in lookup_df.columns:
            lookup_df[column] = pd.to_numeric(lookup_df[column], errors="coerce")
    lookup_df = lookup_df.loc[lookup_df["jumlah_desa"].fillna(0).gt(0)].copy()
    if lookup_df.empty:
        return pd.NA

    low_row = lookup_df.loc[lookup_df["interpretasi_gini"].astype("string").eq(GINI_CATEGORY_ORDER[0])].head(1)
    high_row = lookup_df.loc[lookup_df["interpretasi_gini"].astype("string").eq(GINI_CATEGORY_ORDER[-1])].head(1)
    middle_row = lookup_df.loc[lookup_df["interpretasi_gini"].astype("string").eq(GINI_CATEGORY_ORDER[1])].head(1)

    if not low_row.empty:
        low_upper = low_row["batas_atas"].iloc[0]
        if pd.notna(low_upper) and score <= float(low_upper):
            return GINI_CATEGORY_ORDER[0]

    if not high_row.empty:
        high_lower = high_row["batas_bawah"].iloc[0]
        if pd.notna(high_lower) and score >= float(high_lower):
            return GINI_CATEGORY_ORDER[-1]

    if not middle_row.empty:
        return GINI_CATEGORY_ORDER[1]
    if not high_row.empty:
        return GINI_CATEGORY_ORDER[-1]
    if not low_row.empty:
        return GINI_CATEGORY_ORDER[0]
    return pd.NA


def normalize_household_gini_frame(household_df: pd.DataFrame) -> pd.DataFrame:
    if household_df.empty:
        return pd.DataFrame(columns=["family_id", "nama", "kode_deskel", "deskel", "iid_rumah_tangga"])

    normalized = household_df.copy()
    if "iid_rumah_tangga" not in normalized.columns and "iid_rt" in normalized.columns:
        normalized["iid_rumah_tangga"] = normalized["iid_rt"]
    if "deskel" not in normalized.columns and "deskel_std" in normalized.columns:
        normalized["deskel"] = normalized["deskel_std"]
    if "kode_deskel" not in normalized.columns:
        normalized["kode_deskel"] = pd.NA
    if "deskel" not in normalized.columns:
        normalized["deskel"] = pd.NA
    if "family_id" not in normalized.columns:
        normalized["family_id"] = [f"RT_{idx + 1:06d}" for idx in range(len(normalized))]
    if "nama" not in normalized.columns:
        normalized["nama"] = pd.NA

    if "valid_untuk_indeks" in normalized.columns:
        normalized = normalized.loc[normalized["valid_untuk_indeks"].fillna(False)].copy()

    normalized["iid_rumah_tangga"] = pd.to_numeric(normalized["iid_rumah_tangga"], errors="coerce")
    normalized = normalized.dropna(subset=["iid_rumah_tangga"]).copy()
    normalized = normalized.drop_duplicates(subset=["family_id"], keep="first")
    keep_columns = ["family_id", "nama", "kode_deskel", "deskel", "iid_rumah_tangga"]
    return normalized.loc[:, keep_columns].reset_index(drop=True)


def compute_gini_contribution_components(values: pd.Series | np.ndarray) -> pd.DataFrame:
    array = np.asarray(values, dtype=float)
    valid_mask = np.isfinite(array)
    array = array[valid_mask]
    if array.size == 0:
        return pd.DataFrame(
            columns=[
                "jumlah_selisih_pasangan",
                "kontribusi_gini",
                "porsi_kontribusi_gini",
            ]
        )

    sorted_index = np.argsort(array, kind="mergesort")
    sorted_values = array[sorted_index]
    n = sorted_values.size
    mean_value = float(sorted_values.mean()) if n > 0 else float("nan")
    gini_value = gini_coefficient(sorted_values)

    pairwise_sum = np.zeros(n, dtype=float)
    contribution_abs = np.zeros(n, dtype=float)
    contribution_share = np.zeros(n, dtype=float)

    if n > 1 and np.isfinite(mean_value) and mean_value > 0 and not np.allclose(sorted_values, sorted_values[0]):
        prefix_sum = np.cumsum(sorted_values)
        total_sum = float(prefix_sum[-1])
        index = np.arange(n, dtype=float)
        prefix_before = prefix_sum - sorted_values
        left_component = sorted_values * index - prefix_before
        right_component = (total_sum - prefix_sum) - sorted_values * (n - index - 1)
        pairwise_sum = left_component + right_component
        denominator = 2 * n * n * mean_value
        contribution_abs = pairwise_sum / denominator
        if gini_value > 0:
            contribution_share = contribution_abs / gini_value

    result = pd.DataFrame(
        {
            "jumlah_selisih_pasangan": pairwise_sum,
            "kontribusi_gini": contribution_abs,
            "porsi_kontribusi_gini": contribution_share,
        }
    )
    result.index = sorted_index
    return result.sort_index().reset_index(drop=True)


def build_gini_contributor_table(
    household_df: pd.DataFrame,
    scope_label: str,
    scope_code: object = pd.NA,
    scope_name: object = pd.NA,
) -> pd.DataFrame:
    scope_households = normalize_household_gini_frame(household_df)
    output_columns = [
        "cakupan_analisis",
        "kode_deskel_cakupan",
        "deskel_cakupan",
        "jumlah_kk_cakupan",
        "gini_iid_rumah_tangga_cakupan",
        "interpretasi_gini_cakupan",
        "rata_rata_iid_cakupan",
        "family_id",
        "nama",
        "kode_deskel",
        "deskel",
        "iid_rumah_tangga",
        "deviasi_iid_cakupan",
        "arah_deviasi",
        "jumlah_selisih_pasangan",
        "kontribusi_gini",
        "porsi_kontribusi_gini",
        "peringkat_kontribusi",
    ]
    if scope_households.empty:
        return pd.DataFrame(columns=output_columns)

    mean_value = float(scope_households["iid_rumah_tangga"].mean())
    gini_value = gini_coefficient(scope_households["iid_rumah_tangga"])
    contribution_df = compute_gini_contribution_components(scope_households["iid_rumah_tangga"])

    contributor_df = pd.concat([scope_households.reset_index(drop=True), contribution_df], axis=1)
    contributor_df["cakupan_analisis"] = scope_label
    contributor_df["kode_deskel_cakupan"] = scope_code
    contributor_df["deskel_cakupan"] = scope_name
    contributor_df["jumlah_kk_cakupan"] = int(len(contributor_df))
    contributor_df["gini_iid_rumah_tangga_cakupan"] = float(gini_value)
    contributor_df["interpretasi_gini_cakupan"] = pd.NA
    contributor_df["rata_rata_iid_cakupan"] = mean_value
    contributor_df["deviasi_iid_cakupan"] = contributor_df["iid_rumah_tangga"] - mean_value
    contributor_df["arah_deviasi"] = np.where(
        contributor_df["deviasi_iid_cakupan"] < 0,
        "di bawah rata-rata",
        np.where(
            contributor_df["deviasi_iid_cakupan"] > 0,
            "di atas rata-rata",
            "sama dengan rata-rata",
        ),
    )
    contributor_df = contributor_df.sort_values(
        ["porsi_kontribusi_gini", "kontribusi_gini", "iid_rumah_tangga", "family_id"],
        ascending=[False, False, False, True],
        kind="mergesort",
    ).reset_index(drop=True)
    contributor_df["peringkat_kontribusi"] = np.arange(1, len(contributor_df) + 1)
    return contributor_df.loc[:, output_columns]


def build_gini_summary_row(contributor_df: pd.DataFrame) -> dict[str, object]:
    if contributor_df.empty:
        return {
            "cakupan_analisis": pd.NA,
            "kode_deskel": pd.NA,
            "deskel": pd.NA,
            "jumlah_kk": 0,
            "rata_rata_iid_rumah_tangga": np.nan,
            "gini_iid_rumah_tangga": np.nan,
            "interpretasi_gini": pd.NA,
            "jumlah_kontributor_non_nol": 0,
            "family_id_kontributor_utama": pd.NA,
            "nama_kontributor_utama": pd.NA,
            "iid_kontributor_utama": np.nan,
            "arah_kontributor_utama": pd.NA,
            "porsi_kontributor_utama": np.nan,
        }

    top_row = contributor_df.iloc[0]
    return {
        "cakupan_analisis": top_row["cakupan_analisis"],
        "kode_deskel": top_row["kode_deskel_cakupan"],
        "deskel": top_row["deskel_cakupan"],
        "jumlah_kk": int(top_row["jumlah_kk_cakupan"]),
        "rata_rata_iid_rumah_tangga": float(top_row["rata_rata_iid_cakupan"]),
        "gini_iid_rumah_tangga": float(top_row["gini_iid_rumah_tangga_cakupan"]),
        "interpretasi_gini": pd.NA,
        "jumlah_kontributor_non_nol": int((contributor_df["kontribusi_gini"] > 0).sum()),
        "family_id_kontributor_utama": top_row["family_id"],
        "nama_kontributor_utama": top_row.get("nama", pd.NA),
        "iid_kontributor_utama": float(top_row["iid_rumah_tangga"]),
        "arah_kontributor_utama": top_row["arah_deviasi"],
        "porsi_kontributor_utama": float(top_row["porsi_kontribusi_gini"]),
    }


def build_gini_assessment_tables(household_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    normalized_households = normalize_household_gini_frame(household_df)
    summary_columns = [
        "cakupan_analisis",
        "kode_deskel",
        "deskel",
        "jumlah_kk",
        "rata_rata_iid_rumah_tangga",
        "gini_iid_rumah_tangga",
        "interpretasi_gini",
        "jumlah_kontributor_non_nol",
        "family_id_kontributor_utama",
        "nama_kontributor_utama",
        "iid_kontributor_utama",
        "arah_kontributor_utama",
        "porsi_kontributor_utama",
    ]
    contributor_columns = [
        "cakupan_analisis",
        "kode_deskel_cakupan",
        "deskel_cakupan",
        "jumlah_kk_cakupan",
        "gini_iid_rumah_tangga_cakupan",
        "interpretasi_gini_cakupan",
        "rata_rata_iid_cakupan",
        "family_id",
        "nama",
        "kode_deskel",
        "deskel",
        "iid_rumah_tangga",
        "deviasi_iid_cakupan",
        "arah_deviasi",
        "jumlah_selisih_pasangan",
        "kontribusi_gini",
        "porsi_kontribusi_gini",
        "peringkat_kontribusi",
    ]
    if normalized_households.empty:
        return pd.DataFrame(columns=summary_columns), pd.DataFrame(columns=contributor_columns)

    desa_reference = (
        normalized_households.groupby(["kode_deskel", "deskel"], dropna=False)["iid_rumah_tangga"]
        .apply(gini_coefficient)
        .reset_index(name="gini_iid_rumah_tangga")
    )
    desa_reference, gini_distribution_df = apply_relative_gini_classification(desa_reference)
    desa_category_lookup = desa_reference[["kode_deskel", "deskel", "interpretasi_gini"]].copy()

    contributor_tables: list[pd.DataFrame] = []
    summary_rows: list[dict[str, object]] = []

    overall_name = "Semua desa"
    overall_contributors = build_gini_contributor_table(
        normalized_households,
        scope_label="keseluruhan",
        scope_name=overall_name,
    )
    contributor_tables.append(overall_contributors)
    summary_rows.append(build_gini_summary_row(overall_contributors))

    grouped = normalized_households.groupby(["kode_deskel", "deskel"], dropna=False, sort=True)
    for (kode_deskel, deskel), group_df in grouped:
        desa_name = deskel if pd.notna(deskel) else kode_deskel
        desa_contributors = build_gini_contributor_table(
            group_df,
            scope_label="desa",
            scope_code=kode_deskel,
            scope_name=desa_name,
        )
        contributor_tables.append(desa_contributors)
        summary_rows.append(build_gini_summary_row(desa_contributors))

    summary_df = pd.DataFrame(summary_rows)
    desa_summary_mask = summary_df["cakupan_analisis"].astype("string").eq("desa")
    if desa_summary_mask.any():
        summary_df.loc[desa_summary_mask, "kode_deskel"] = summary_df.loc[desa_summary_mask, "kode_deskel"].astype("string")
        desa_category_lookup["kode_deskel"] = desa_category_lookup["kode_deskel"].astype("string")
        summary_df = summary_df.merge(
            desa_category_lookup.rename(columns={"interpretasi_gini": "interpretasi_gini_relatif"}),
            on=["kode_deskel", "deskel"],
            how="left",
        )
        summary_df["interpretasi_gini"] = summary_df["interpretasi_gini_relatif"].combine_first(summary_df["interpretasi_gini"])
        summary_df = summary_df.drop(columns=["interpretasi_gini_relatif"])

    overall_mask = summary_df["cakupan_analisis"].astype("string").eq("keseluruhan")
    if overall_mask.any():
        overall_gini_value = summary_df.loc[overall_mask, "gini_iid_rumah_tangga"].iloc[0]
        summary_df.loc[overall_mask, "interpretasi_gini"] = interpret_gini_value(
            overall_gini_value,
            gini_distribution_df,
        )

    summary_df["_cakupan_order"] = np.where(summary_df["cakupan_analisis"].eq("keseluruhan"), 0, 1)
    summary_df = summary_df.sort_values(
        ["_cakupan_order", "gini_iid_rumah_tangga", "jumlah_kk", "deskel"],
        ascending=[True, False, False, True],
        kind="mergesort",
    ).drop(columns="_cakupan_order").reset_index(drop=True)

    contributor_df = pd.concat(contributor_tables, ignore_index=True, sort=False)
    desa_contributor_mask = contributor_df["cakupan_analisis"].astype("string").eq("desa")
    if desa_contributor_mask.any():
        contributor_df.loc[desa_contributor_mask, "kode_deskel_cakupan"] = contributor_df.loc[
            desa_contributor_mask,
            "kode_deskel_cakupan",
        ].astype("string")
        desa_contributor_lookup = desa_category_lookup.rename(
            columns={
                "kode_deskel": "kode_deskel_cakupan",
                "deskel": "deskel_cakupan",
                "interpretasi_gini": "interpretasi_gini_relatif",
            }
        )
        contributor_df = contributor_df.merge(
            desa_contributor_lookup,
            on=["kode_deskel_cakupan", "deskel_cakupan"],
            how="left",
        )
        contributor_df["interpretasi_gini_cakupan"] = contributor_df["interpretasi_gini_relatif"].combine_first(
            contributor_df["interpretasi_gini_cakupan"]
        )
        contributor_df = contributor_df.drop(columns=["interpretasi_gini_relatif"])

    overall_contributor_mask = contributor_df["cakupan_analisis"].astype("string").eq("keseluruhan")
    if overall_contributor_mask.any():
        overall_category = summary_df.loc[
            summary_df["cakupan_analisis"].astype("string").eq("keseluruhan"),
            "interpretasi_gini",
        ].iloc[0]
        contributor_df.loc[overall_contributor_mask, "interpretasi_gini_cakupan"] = overall_category

    contributor_df["_cakupan_order"] = np.where(contributor_df["cakupan_analisis"].eq("keseluruhan"), 0, 1)
    contributor_df = contributor_df.sort_values(
        ["_cakupan_order", "deskel_cakupan", "peringkat_kontribusi"],
        ascending=[True, True, True],
        kind="mergesort",
    ).drop(columns="_cakupan_order").reset_index(drop=True)
    return summary_df.loc[:, summary_columns], contributor_df.loc[:, contributor_columns]


def compute_dimension_mean(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    return df[columns].mean(axis=1, skipna=True)


def build_household_index(
    person_df: pd.DataFrame,
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
    missing_threshold: float = MISSING_THRESHOLD,
    phone_winsor_quantile: float = PHONE_WINSOR_QUANTILE,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if school_age_min > school_age_max:
        raise ValueError("Batas bawah usia sekolah tidak boleh lebih besar dari batas atas.")

    df = person_df.copy()
    df["family_id"] = derive_family_id(df)
    df = df[df["family_id"].notna()].copy()

    df["subjek_norm"] = normalize_text_series(df["subjek"])
    df["status_dalam_keluarga_norm"] = normalize_text_series(df["status_dalam_keluarga"])
    df["is_head"] = build_head_mask(df)
    df["usia_num"] = pd.to_numeric(df["usia"], errors="coerce")
    df["partisipasi_sekolah_norm"] = normalize_text_series(df["partisipasi_sekolah"])
    df["jumlah_isian_organisasi"] = df.apply(
        lambda row: count_combined_multivalue_items(row["par_organisasi"], row["organisasi_nama"]),
        axis=1,
    )
    df["jumlah_isian_masyarakat"] = df["par_masyarakat"].apply(count_multivalue_items)
    df["jumlah_kebijakan_ya"] = normalize_text_series(df["par_kebijakan"]).isin(YES_VALUES).astype(int)

    df_sorted = df.sort_values(["family_id", "is_head"], ascending=[True, False], kind="mergesort")
    preferred_row = df_sorted.drop_duplicates(subset="family_id", keep="first").set_index("family_id")

    household_size_counted = df.groupby("family_id").size().rename("jumlah_anggota_rumah_tangga_tercatat")
    head_count = df.groupby("family_id")["is_head"].sum().rename("jumlah_kepala_keluarga_tercatat")
    has_head = head_count.gt(0).rename("ada_kepala_keluarga")
    household_size_reported = pd.to_numeric(preferred_row["jml_keluarga"], errors="coerce").clip(lower=0)
    household_size_for_scoring = household_size_reported.reindex(preferred_row.index)
    household_size_for_scoring = household_size_for_scoring.where(household_size_for_scoring.gt(0), np.nan)
    household_size_for_scoring = household_size_for_scoring.fillna(
        household_size_counted.reindex(preferred_row.index, fill_value=0).astype(float)
    )

    eligible_school_mask = school_age_mask(df["usia_num"], school_age_min, school_age_max)
    known_school_mask = eligible_school_mask & df["partisipasi_sekolah_norm"].ne("")
    in_school_mask = eligible_school_mask & df["partisipasi_sekolah_norm"].eq("sedang sekolah")

    school_total = (
        df.loc[eligible_school_mask, "family_id"]
        .value_counts(sort=False)
        .rename("jumlah_anggota_usia_sekolah")
        .reindex(preferred_row.index, fill_value=0)
    )
    school_known = (
        df.loc[known_school_mask, "family_id"]
        .value_counts(sort=False)
        .rename("jumlah_status_sekolah_terisi")
        .reindex(preferred_row.index, fill_value=0)
    )
    school_in_school = (
        df.loc[in_school_mask, "family_id"]
        .value_counts(sort=False)
        .rename("jumlah_anggota_sedang_sekolah")
        .reindex(preferred_row.index, fill_value=0)
    )
    wifi_items = df.groupby("family_id", dropna=False)["wifi"].apply(collect_multivalue_items).reindex(preferred_row.index)
    provider_items = (
        df.groupby("family_id", dropna=False)["hp_provider"].apply(collect_multivalue_items).reindex(preferred_row.index)
    )
    medsos_items = df.groupby("family_id", dropna=False)["medsos"].apply(collect_multivalue_items).reindex(preferred_row.index)
    media_informasi_items = (
        df.groupby("family_id", dropna=False)["media_informasi"].apply(collect_multivalue_items).reindex(preferred_row.index)
    )
    komunikasi_spend = (
        df.groupby("family_id", dropna=False)["rp_komunikasi"]
        .apply(lambda series: pd.to_numeric(series, errors="coerce").max())
        .reindex(preferred_row.index)
    )
    kebijakan_yes_count = (
        df.groupby("family_id", dropna=False)["jumlah_kebijakan_ya"]
        .sum()
        .reindex(preferred_row.index, fill_value=0)
        .astype(int)
    )
    organisasi_kepala_count = (
        df.loc[df["is_head"]]
        .groupby("family_id")["jumlah_isian_organisasi"]
        .sum()
        .reindex(preferred_row.index, fill_value=0)
        .astype(int)
    )
    organisasi_anggota_count = (
        df.loc[~df["is_head"]]
        .groupby("family_id")["jumlah_isian_organisasi"]
        .sum()
        .reindex(preferred_row.index, fill_value=0)
        .astype(int)
    )
    masyarakat_kepala_count = (
        df.loc[df["is_head"]]
        .groupby("family_id")["jumlah_isian_masyarakat"]
        .sum()
        .reindex(preferred_row.index, fill_value=0)
        .astype(int)
    )
    masyarakat_anggota_count = (
        df.loc[~df["is_head"]]
        .groupby("family_id")["jumlah_isian_masyarakat"]
        .sum()
        .reindex(preferred_row.index, fill_value=0)
        .astype(int)
    )

    household_df = pd.DataFrame(index=preferred_row.index)
    base_columns = [
        "abs_id",
        "subjek",
        "kode_bangunan",
        "kode_deskel",
        "deskel",
        "dusun",
        "rw",
        "lat",
        "lng",
        "nama",
        "suku",
        "jml_keluarga",
        "hp_punya",
        "hp_jumlah",
        "elektronik_rumah",
        "wifi",
        "hp_provider",
        "rp_komunikasi",
        "status_dalam_keluarga",
        "ijazah",
        "par_organisasi",
        "organisasi_nama",
        "par_masyarakat",
        "medsos",
        "media_informasi",
        "par_kebijakan",
    ]
    for column in base_columns:
        household_df[column] = preferred_row[column] if column in preferred_row.columns else pd.NA

    household_df["jumlah_anggota_rumah_tangga"] = household_size_for_scoring.round().astype(int)
    household_df["jumlah_anggota_rumah_tangga_tercatat"] = (
        household_size_counted.reindex(household_df.index, fill_value=0).astype(int)
    )
    household_df["jumlah_kepala_keluarga_tercatat"] = head_count.reindex(household_df.index, fill_value=0).astype(int)
    household_df["ada_kepala_keluarga"] = has_head.reindex(household_df.index, fill_value=False).astype(bool)
    household_df["jumlah_anggota_usia_sekolah"] = school_total.astype(int)
    household_df["jumlah_status_sekolah_terisi"] = school_known.astype(int)
    household_df["jumlah_anggota_sedang_sekolah"] = school_in_school.astype(int)

    household_df["deskel"] = household_df["deskel"].astype("string").str.strip()
    household_df["deskel_std"] = normalize_text_series(household_df["deskel"]).replace({"": pd.NA})
    household_df["dusun"] = household_df["dusun"].astype("string").str.strip()
    household_df["dusun_std"] = standardize_dusun_label(household_df["dusun"])
    household_df["rw"] = household_df["rw"].astype("string").str.strip()

    household_df["hp_jumlah_num"] = pd.to_numeric(household_df["hp_jumlah"], errors="coerce")
    hp_owned_norm = normalize_text_series(household_df["hp_punya"])
    household_df["hp_jumlah_terstandar"] = household_df["hp_jumlah_num"].clip(lower=0)
    household_df.loc[
        hp_owned_norm.isin(YES_VALUES) & household_df["hp_jumlah_terstandar"].isna(),
        "hp_jumlah_terstandar",
    ] = 1.0
    household_df.loc[
        hp_owned_norm.isin(NO_VALUES) & household_df["hp_jumlah_terstandar"].isna(),
        "hp_jumlah_terstandar",
    ] = 0.0
    household_df["flag_hp_tidak_konsisten"] = (
        hp_owned_norm.isin(NO_VALUES) & household_df["hp_jumlah_terstandar"].fillna(0).gt(0)
    )
    household_df["indikator_hp_dimiliki"] = household_df["hp_jumlah_terstandar"].apply(score_binary_presence)

    household_df["rasio_hp_terhadap_anggota"] = (
        household_df["hp_jumlah_terstandar"] / household_df["jumlah_anggota_rumah_tangga"].replace(0, np.nan)
    ).clip(lower=0, upper=1)
    household_df["indikator_kecukupan_hp"] = household_df["rasio_hp_terhadap_anggota"].astype(float)

    household_df["jumlah_perangkat_produktif_rumah_tangga"] = household_df["elektronik_rumah"].apply(
        lambda value: count_keyword_matches(value, DIGITAL_PRODUCTIVE_DEVICE_KEYWORDS)
    )
    household_df["indikator_perangkat_produktif"] = household_df["jumlah_perangkat_produktif_rumah_tangga"].apply(
        score_binary_presence
    )

    household_df["wifi_teragregasi"] = wifi_items.apply(lambda values: ",".join(values) if values else pd.NA)
    household_df["hp_provider_teragregasi"] = provider_items.apply(lambda values: ",".join(values) if values else pd.NA)
    household_df["rp_komunikasi_tertinggi"] = komunikasi_spend
    household_df["indikator_akses_internet"] = household_df.apply(
        lambda row: score_internet_access(
            row["wifi_teragregasi"],
            row["hp_provider_teragregasi"],
            row["rp_komunikasi_tertinggi"],
        ),
        axis=1,
    ).astype(float)
    household_df["indikator_pendidikan_kepala"] = household_df["ijazah"].apply(score_education)
    household_df["rasio_sekolah_berlaku"] = household_df["jumlah_anggota_usia_sekolah"].gt(0)
    household_df["indikator_rasio_sekolah"] = np.nan

    status_sekolah_lengkap_mask = (
        household_df["jumlah_anggota_usia_sekolah"].eq(0)
        | household_df["jumlah_status_sekolah_terisi"].eq(household_df["jumlah_anggota_usia_sekolah"])
    )
    valid_school_ratio_mask = household_df["rasio_sekolah_berlaku"] & status_sekolah_lengkap_mask
    school_ratio = (
        household_df["jumlah_anggota_sedang_sekolah"] / household_df["jumlah_anggota_usia_sekolah"].replace(0, np.nan)
    ).clip(lower=0, upper=1)
    household_df.loc[valid_school_ratio_mask, "indikator_rasio_sekolah"] = school_ratio.loc[valid_school_ratio_mask]

    household_df["jumlah_organisasi_kepala"] = organisasi_kepala_count
    household_df["jumlah_organisasi_anggota"] = organisasi_anggota_count
    household_df["jumlah_partisipasi_masyarakat_kepala"] = masyarakat_kepala_count
    household_df["jumlah_partisipasi_masyarakat_anggota"] = masyarakat_anggota_count
    household_df["indikator_organisasi_kepala"] = household_df["jumlah_organisasi_kepala"].apply(score_zero_one_many)
    household_df["indikator_organisasi_anggota"] = household_df["jumlah_organisasi_anggota"].apply(score_zero_one_many)
    household_df["indikator_partisipasi_masyarakat_kepala"] = household_df["jumlah_partisipasi_masyarakat_kepala"].apply(
        score_zero_one_many
    )
    household_df["indikator_partisipasi_masyarakat_anggota"] = household_df["jumlah_partisipasi_masyarakat_anggota"].apply(
        score_zero_one_many
    )
    household_df["medsos_teragregasi"] = medsos_items.apply(lambda values: ",".join(values) if values else pd.NA)
    household_df["media_informasi_teragregasi"] = media_informasi_items.apply(
        lambda values: ",".join(values) if values else pd.NA
    )
    household_df["jumlah_partisipasi_kebijakan"] = kebijakan_yes_count
    household_df["indikator_medsos"] = household_df["medsos_teragregasi"].apply(score_social_media_use)
    household_df["indikator_media_informasi"] = household_df["media_informasi_teragregasi"].apply(score_media_information)
    household_df["indikator_partisipasi_kebijakan"] = household_df["jumlah_partisipasi_kebijakan"].apply(score_zero_one_many)

    scored_indicator_cols = [
        "indikator_hp_dimiliki",
        "indikator_kecukupan_hp",
        "indikator_perangkat_produktif",
        "indikator_akses_internet",
        "indikator_pendidikan_kepala",
        "indikator_organisasi_kepala",
        "indikator_organisasi_anggota",
        "indikator_partisipasi_masyarakat_kepala",
        "indikator_partisipasi_masyarakat_anggota",
        "indikator_medsos",
        "indikator_media_informasi",
        "indikator_partisipasi_kebijakan",
    ]
    core_indicator_cols = scored_indicator_cols + ["indikator_rasio_sekolah"]

    applicable_indicator_count = pd.Series(len(core_indicator_cols), index=household_df.index, dtype=float)
    applicable_indicator_count = applicable_indicator_count - (~household_df["rasio_sekolah_berlaku"]).astype(int)

    missing_core_count = household_df[scored_indicator_cols].isna().sum(axis=1)
    missing_core_count = missing_core_count + (
        household_df["indikator_rasio_sekolah"].isna() & household_df["rasio_sekolah_berlaku"]
    ).astype(int)

    household_df["proporsi_indikator_inti_hilang"] = (missing_core_count / applicable_indicator_count.replace(0, np.nan)).fillna(0)
    household_df["lokasi_tidak_lengkap"] = household_df["deskel_std"].fillna("").eq("")
    household_df["lolos_kelengkapan_inti"] = household_df["proporsi_indikator_inti_hilang"].le(missing_threshold)
    household_df["valid_untuk_indeks"] = (
        household_df["ada_kepala_keluarga"]
        & household_df["lolos_kelengkapan_inti"]
        & ~household_df["lokasi_tidak_lengkap"]
    )

    excluded_df = household_df.loc[~household_df["valid_untuk_indeks"]].copy()
    if not excluded_df.empty:
        exclusion_reason = []
        for _, row in excluded_df.iterrows():
            reasons: list[str] = []
            if not bool(row["ada_kepala_keluarga"]):
                reasons.append("tanpa_kepala_keluarga")
            if bool(row["lokasi_tidak_lengkap"]):
                reasons.append("lokasi_deskel_tidak_lengkap")
            if float(row["proporsi_indikator_inti_hilang"]) > missing_threshold:
                reasons.append("indikator_inti_hilang_lebih_dari_20_persen")
            exclusion_reason.append(";".join(reasons) if reasons else "tidak_memenuhi_kriteria")
        excluded_df["alasan_dikeluarkan"] = exclusion_reason

    valid_df = household_df.loc[household_df["valid_untuk_indeks"]].copy()

    imputation_summary: list[dict[str, object]] = []
    for column in scored_indicator_cols:
        missing_before = int(valid_df[column].isna().sum())
        fill_value = float(valid_df[column].mode(dropna=True).iloc[0]) if valid_df[column].notna().any() else 0.0
        valid_df[column] = valid_df[column].fillna(fill_value)
        imputation_summary.append(
            {"indikator": column, "tipe_imputasi": "modus", "nilai_imputasi": fill_value, "jumlah_diimputasi": missing_before}
        )

    school_missing_before = int((valid_df["indikator_rasio_sekolah"].isna() & valid_df["rasio_sekolah_berlaku"]).sum())
    school_fill_value = (
        float(valid_df.loc[valid_df["rasio_sekolah_berlaku"], "indikator_rasio_sekolah"].mode(dropna=True).iloc[0])
        if valid_df.loc[valid_df["rasio_sekolah_berlaku"], "indikator_rasio_sekolah"].notna().any()
        else 0.0
    )
    applicable_school_mask = valid_df["rasio_sekolah_berlaku"] & valid_df["indikator_rasio_sekolah"].isna()
    valid_df.loc[applicable_school_mask, "indikator_rasio_sekolah"] = school_fill_value
    imputation_summary.append(
        {
            "indikator": "indikator_rasio_sekolah",
            "tipe_imputasi": "modus_pada_rumah_tangga_yang_berlaku",
            "nilai_imputasi": school_fill_value,
            "jumlah_diimputasi": school_missing_before,
        }
    )

    valid_df["dimensi_akses_perangkat"] = compute_dimension_mean(
        valid_df,
        ["indikator_hp_dimiliki", "indikator_kecukupan_hp", "indikator_perangkat_produktif"],
    )
    valid_df["dimensi_konektivitas"] = valid_df["indikator_akses_internet"]
    valid_df["dimensi_kapasitas_manusia"] = compute_dimension_mean(
        valid_df,
        ["indikator_pendidikan_kepala", "indikator_rasio_sekolah"],
    )
    valid_df["dimensi_penggunaan_digital"] = compute_dimension_mean(
        valid_df,
        ["indikator_medsos", "indikator_media_informasi", "indikator_partisipasi_kebijakan"],
    )
    valid_df["dimensi_lingkungan_sosial"] = compute_dimension_mean(
        valid_df,
        [
            "indikator_organisasi_kepala",
            "indikator_organisasi_anggota",
            "indikator_partisipasi_masyarakat_kepala",
            "indikator_partisipasi_masyarakat_anggota",
        ],
    )

    valid_df["iid_rt"] = (
        DIMENSION_WEIGHTS["akses_perangkat"] * valid_df["dimensi_akses_perangkat"]
        + DIMENSION_WEIGHTS["konektivitas"] * valid_df["dimensi_konektivitas"]
        + DIMENSION_WEIGHTS["kapasitas_manusia"] * valid_df["dimensi_kapasitas_manusia"]
        + DIMENSION_WEIGHTS["penggunaan_digital"] * valid_df["dimensi_penggunaan_digital"]
        + DIMENSION_WEIGHTS["lingkungan_sosial"] * valid_df["dimensi_lingkungan_sosial"]
    ).clip(lower=0, upper=1)

    valid_df["ikd_rt"] = 1 - valid_df["iid_rt"]
    valid_df = valid_df.reset_index().rename(columns={"index": "family_id"})
    excluded_df = excluded_df.reset_index().rename(columns={"index": "family_id"})

    processing_summary = pd.DataFrame(
        [
            {"metrik": "jumlah_baris_sumber", "nilai": int(len(df))},
            {"metrik": "jumlah_rumah_tangga_teridentifikasi", "nilai": int(household_df.shape[0])},
            {"metrik": "jumlah_rumah_tangga_dengan_kepala", "nilai": int(household_df["ada_kepala_keluarga"].sum())},
            {"metrik": "jumlah_rumah_tangga_valid", "nilai": int(valid_df.shape[0])},
            {"metrik": "jumlah_rumah_tangga_dikeluarkan", "nilai": int(excluded_df.shape[0])},
            {"metrik": "batas_usia_sekolah_min", "nilai": school_age_min},
            {"metrik": "batas_usia_sekolah_max", "nilai": school_age_max},
            {
                "metrik": "cakupan_usia_partisipasi_sekolah",
                "nilai": (
                    build_school_stage_summary()
                    if is_default_school_age_range(school_age_min, school_age_max)
                    else build_school_age_range_text(school_age_min, school_age_max)
                ),
            },
        ]
        + imputation_summary
    )

    return valid_df, excluded_df, processing_summary


def _mode_or_na(series: pd.Series) -> object:
    clean = series.dropna().astype(str).str.strip()
    if clean.empty:
        return pd.NA
    return clean.mode().iat[0]


def build_household_master(
    valid_households: pd.DataFrame,
    excluded_households: pd.DataFrame,
    iid_classifier: object = classify_iid_rt,
) -> pd.DataFrame:
    household_master = pd.concat([valid_households, excluded_households], ignore_index=True, sort=False)
    for column in [*DIMENSION_OUTPUT_MAP.keys(), "iid_rt", "ikd_rt"]:
        if column not in household_master.columns:
            household_master[column] = np.nan
    if "alasan_dikeluarkan" not in household_master.columns:
        household_master["alasan_dikeluarkan"] = pd.NA
    household_master["status_indeks_rt"] = np.where(
        household_master["valid_untuk_indeks"].fillna(False),
        "valid",
        "dikeluarkan",
    )
    household_master["iid_rumah_tangga"] = household_master["iid_rt"]
    household_master["kategori_iid_rt"] = household_master["iid_rumah_tangga"].apply(iid_classifier)
    household_master = household_master.sort_values(
        ["kode_deskel", "deskel_std", "dusun_std", "family_id"],
        kind="mergesort",
    ).reset_index(drop=True)
    return household_master


def build_keluarga_output(
    person_df: pd.DataFrame,
    household_master: pd.DataFrame,
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> pd.DataFrame:
    keluarga_df = person_df.copy()
    keluarga_df["family_id"] = derive_family_id(keluarga_df)
    keluarga_df["subjek_norm"] = normalize_text_series(keluarga_df["subjek"])
    keluarga_df["is_kepala_keluarga"] = build_head_mask(keluarga_df)
    keluarga_df["usia_num"] = pd.to_numeric(keluarga_df["usia"], errors="coerce")
    keluarga_df["anggota_usia_sekolah"] = school_age_mask(
        keluarga_df["usia_num"],
        school_age_min,
        school_age_max,
    )

    household_columns = [
        "family_id",
        "valid_untuk_indeks",
        *INDICATOR_OUTPUT_MAP.keys(),
        *DIMENSION_OUTPUT_MAP.keys(),
        "iid_rumah_tangga",
        "kategori_iid_rt",
    ]
    keluarga_df = keluarga_df.merge(household_master[household_columns], on="family_id", how="left")
    keluarga_df = keluarga_df.loc[keluarga_df["valid_untuk_indeks"].eq(True)].copy()

    head_only_indicator_columns = [
        "indikator_hp_dimiliki",
        "indikator_kecukupan_hp",
        "indikator_perangkat_produktif",
        "indikator_akses_internet",
        "indikator_pendidikan_kepala",
        "indikator_organisasi_kepala",
        "indikator_organisasi_anggota",
        "indikator_partisipasi_masyarakat_kepala",
        "indikator_partisipasi_masyarakat_anggota",
        "indikator_medsos",
        "indikator_media_informasi",
        "indikator_partisipasi_kebijakan",
    ]
    for column in head_only_indicator_columns:
        keluarga_df[column] = keluarga_df[column].where(keluarga_df["is_kepala_keluarga"], pd.NA)

    keluarga_df["indikator_rasio_sekolah"] = keluarga_df["indikator_rasio_sekolah"].where(
        keluarga_df["anggota_usia_sekolah"],
        pd.NA,
    )

    head_only_columns = [*DIMENSION_OUTPUT_MAP.keys(), "iid_rumah_tangga", "kategori_iid_rt"]
    for column in head_only_columns:
        keluarga_df[column] = keluarga_df[column].where(keluarga_df["is_kepala_keluarga"], pd.NA)

    keluarga_df = keluarga_df.rename(columns={**INDICATOR_OUTPUT_MAP, **DIMENSION_OUTPUT_MAP, "lng": "long"})

    ordered_columns = [
        "family_id",
        "abs_id",
        "subjek",
        "nama",
        "usia",
        "lat",
        "long",
        "kode_deskel",
        "deskel",
        *INDICATOR_OUTPUT_MAP.values(),
        *DIMENSION_OUTPUT_MAP.values(),
        "iid_rumah_tangga",
        "kategori_iid_rt",
    ]
    ordered_columns = [col for col in ordered_columns if col in keluarga_df.columns]
    keluarga_df = keluarga_df.loc[:, ordered_columns]
    keluarga_df = keluarga_df.assign(_head_sort=keluarga_df["subjek"].astype("string").str.lower().eq("kepala keluarga"))
    keluarga_df = keluarga_df.sort_values(
        ["kode_deskel", "family_id", "_head_sort", "abs_id"],
        ascending=[True, True, False, True],
        kind="mergesort",
    ).drop(columns=["_head_sort"]).reset_index(drop=True)
    keluarga_df = keluarga_df.drop(columns=["subjek_norm", "usia_num", "anggota_usia_sekolah", "valid_untuk_indeks"], errors="ignore")
    return keluarga_df


def build_desa_summary(household_master: pd.DataFrame) -> pd.DataFrame:
    group_keys = ["kode_deskel", "deskel_std"]
    valid_households = household_master.loc[household_master["valid_untuk_indeks"]].copy()

    count_summary = valid_households.groupby(group_keys, dropna=False).agg(
        deskel=("deskel", _mode_or_na),
        jumlah_kk=("family_id", "count"),
    ).reset_index()

    agg_spec: dict[str, tuple[str, object]] = {
        "iid_desa": ("iid_rumah_tangga", "mean"),
    }
    for column in INDICATOR_OUTPUT_MAP:
        agg_spec[column] = (column, "mean")
    for column in DIMENSION_OUTPUT_MAP:
        agg_spec[column] = (column, "mean")

    desa_summary = valid_households.groupby(group_keys, dropna=False).agg(**agg_spec).reset_index()
    gini_summary = (
        valid_households.groupby(group_keys, dropna=False)["iid_rumah_tangga"]
        .apply(gini_coefficient)
        .reset_index(name="gini_iid_rumah_tangga")
    )
    desa_summary = desa_summary.merge(count_summary, on=group_keys, how="left")
    desa_summary = desa_summary.merge(gini_summary, on=group_keys, how="left")
    desa_summary["iid_desa"] = desa_summary["iid_desa"].clip(lower=0, upper=1)
    desa_summary["ikd_desa"] = 1 - desa_summary["iid_desa"]
    desa_summary["ikd_desa"] = desa_summary["ikd_desa"].clip(lower=0, upper=1)
    desa_summary["gini_iid_rumah_tangga"] = desa_summary["gini_iid_rumah_tangga"].clip(lower=0, upper=1)
    desa_summary, _ = apply_relative_gini_classification(desa_summary)

    rename_map = {
        **INDICATOR_OUTPUT_MAP,
        **DIMENSION_OUTPUT_MAP,
    }
    desa_summary = desa_summary.rename(columns=rename_map)
    desa_summary = desa_summary.drop(columns=["deskel_std"])

    ordered_columns = [
        "kode_deskel",
        "deskel",
        "jumlah_kk",
        *INDICATOR_OUTPUT_MAP.values(),
        *DIMENSION_OUTPUT_MAP.values(),
        "iid_desa",
        "ikd_desa",
        "gini_iid_rumah_tangga",
        "interpretasi_gini",
    ]
    ordered_columns = [col for col in ordered_columns if col in desa_summary.columns]
    desa_summary = desa_summary.loc[:, ordered_columns]
    desa_summary = desa_summary.sort_values(["kode_deskel"], kind="mergesort").reset_index(drop=True)
    return desa_summary


def _build_advanced_analysis_empty_tables() -> dict[str, pd.DataFrame]:
    return {
        "analisis_determinasi_dimensi": pd.DataFrame(columns=["Dimensi", "R2 IID Desa"]),
        "analisis_determinasi_variabel": pd.DataFrame(columns=["Dimensi", "Variabel", "R2 Dimensi", "R2 IID Desa"]),
        "analisis_sensitivitas_oat": pd.DataFrame(
            columns=[
                "Dimensi",
                "Rata-rata Delta IID Desa",
                "Rata-rata Kenaikan IID Desa (%)",
                "Rata-rata Delta Deprivasi Digital",
                "Rata-rata Penurunan Deprivasi Digital (%)",
                "Skenario OAT",
            ]
        ),
        "analisis_shapley_variabel": pd.DataFrame(
            columns=[
                "Dimensi",
                "Variabel",
                "Shapley R2 Dimensi",
                "Proporsi Shapley Dimensi",
                "Shapley R2 IID Desa",
                "Proporsi Shapley IID Desa",
            ]
        ),
    }


def _resolve_analysis_label_lookup(variable_explanation: pd.DataFrame) -> dict[str, str]:
    lookup = (
        variable_explanation.loc[:, ["nama_variabel", "label_konsep"]]
        .dropna(subset=["nama_variabel"])
        .drop_duplicates(subset=["nama_variabel"], keep="first")
        .set_index("nama_variabel")["label_konsep"]
        .astype(str)
        .to_dict()
        if not variable_explanation.empty and {"nama_variabel", "label_konsep"}.issubset(variable_explanation.columns)
        else {}
    )
    if not lookup:
        fallback_explanation = build_variable_explanation()
        lookup = (
            fallback_explanation.loc[:, ["nama_variabel", "label_konsep"]]
            .set_index("nama_variabel")["label_konsep"]
            .astype(str)
            .to_dict()
        )
    return lookup


def _build_dimension_display_name(dimension_column: str, dimension_code: str, label_lookup: dict[str, str]) -> str:
    label = label_lookup.get(dimension_column, dimension_column)
    return f"{label} ({dimension_code})"


def _build_indicator_display_name(indicator_column: str, indicator_code: str, label_lookup: dict[str, str]) -> str:
    label = label_lookup.get(indicator_column, indicator_column)
    return f"{label} ({indicator_code})"


def build_oat_sensitivity_table(
    desa_summary: pd.DataFrame,
    variable_explanation: pd.DataFrame,
    increment_value: float = OAT_DIMENSION_INCREMENT,
) -> pd.DataFrame:
    oat_empty = _build_advanced_analysis_empty_tables()["analisis_sensitivitas_oat"]
    if desa_summary.empty:
        return oat_empty.copy()

    try:
        increment_value = float(increment_value)
    except (TypeError, ValueError):
        increment_value = OAT_DIMENSION_INCREMENT
    increment_value = float(np.clip(increment_value, 0.0, 1.0))

    label_lookup = _resolve_analysis_label_lookup(variable_explanation)
    numeric_df = desa_summary.copy()
    core_columns = {"iid_desa", "ikd_desa"}
    for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
        core_columns.add(spec["dimension_column"])

    available_columns = [column for column in core_columns if column in numeric_df.columns]
    for column in available_columns:
        numeric_df[column] = pd.to_numeric(numeric_df[column], errors="coerce")

    dimension_columns = [
        spec["dimension_column"]
        for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS
        if spec["dimension_column"] in numeric_df.columns
    ]
    if not dimension_columns:
        return oat_empty.copy()

    if "iid_desa" in numeric_df.columns and numeric_df["iid_desa"].notna().any():
        base_iid_series = pd.to_numeric(numeric_df["iid_desa"], errors="coerce")
    else:
        base_iid_series = pd.Series(0.0, index=numeric_df.index, dtype=float)
        for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
            dimension_column = spec["dimension_column"]
            if dimension_column in numeric_df.columns:
                base_iid_series = base_iid_series + (numeric_df[dimension_column].fillna(0.0) * float(spec["weight"]))
        base_iid_series = base_iid_series.clip(lower=0.0, upper=1.0)
        numeric_df["iid_desa"] = base_iid_series

    if "ikd_desa" in numeric_df.columns:
        base_ikd_series = pd.to_numeric(numeric_df["ikd_desa"], errors="coerce")
        base_ikd_series = base_ikd_series.where(base_ikd_series.notna(), 1.0 - base_iid_series)
    else:
        base_ikd_series = 1.0 - base_iid_series
        numeric_df["ikd_desa"] = base_ikd_series

    oat_rows: list[dict[str, object]] = []
    increment_percent = increment_value * 100.0
    increment_percent_label = format_numeric_label(increment_percent)

    for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
        dimension_column = spec["dimension_column"]
        if dimension_column not in dimension_columns:
            continue

        dimension_name = _build_dimension_display_name(dimension_column, spec["dimension_code"], label_lookup)
        simulated_dimensions = numeric_df.loc[:, dimension_columns].copy()
        simulated_dimensions[dimension_column] = simulated_dimensions[dimension_column].add(increment_value).clip(upper=1.0)
        simulated_iid_series = pd.Series(0.0, index=simulated_dimensions.index, dtype=float)
        for weight_spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
            weighted_dimension = weight_spec["dimension_column"]
            if weighted_dimension in simulated_dimensions.columns:
                simulated_iid_series = simulated_iid_series + (
                    simulated_dimensions[weighted_dimension].fillna(0.0) * float(weight_spec["weight"])
                )
        simulated_iid_series = simulated_iid_series.clip(lower=0.0, upper=1.0)
        base_deprivation_series = base_ikd_series.clip(lower=0.0, upper=1.0)
        simulated_deprivation_series = (1.0 - simulated_iid_series).clip(lower=0.0, upper=1.0)

        iid_delta = simulated_iid_series - base_iid_series
        deprivation_delta = simulated_deprivation_series - base_deprivation_series

        iid_increase_percent = pd.Series(np.nan, index=simulated_iid_series.index, dtype=float)
        iid_valid_mask = base_iid_series > 0
        iid_increase_percent.loc[iid_valid_mask] = (
            (simulated_iid_series.loc[iid_valid_mask] - base_iid_series.loc[iid_valid_mask])
            / base_iid_series.loc[iid_valid_mask]
        ) * 100.0
        iid_increase_percent.loc[base_iid_series.eq(0) & simulated_iid_series.eq(0)] = 0.0

        deprivation_decrease_percent = pd.Series(np.nan, index=simulated_deprivation_series.index, dtype=float)
        deprivation_valid_mask = base_deprivation_series > 0
        deprivation_decrease_percent.loc[deprivation_valid_mask] = (
            (base_deprivation_series.loc[deprivation_valid_mask] - simulated_deprivation_series.loc[deprivation_valid_mask])
            / base_deprivation_series.loc[deprivation_valid_mask]
        ) * 100.0
        deprivation_decrease_percent.loc[
            base_deprivation_series.eq(0) & simulated_deprivation_series.eq(0)
        ] = 0.0

        oat_rows.append(
            {
                "Dimensi": dimension_name,
                "Rata-rata Delta IID Desa": float(
                    iid_delta.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Rata-rata Kenaikan IID Desa (%)": float(
                    iid_increase_percent.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Rata-rata Delta Deprivasi Digital": float(
                    deprivation_delta.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Rata-rata Penurunan Deprivasi Digital (%)": float(
                    deprivation_decrease_percent.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Skenario OAT": (
                    f"+{format_numeric_label(increment_value)} ({increment_percent_label} persen) pada dimensi (skala 0-1), "
                    "lalu dihitung perubahan IID Desa dan deprivasi digital sebagai 1 - IID"
                ),
            }
        )

    return pd.DataFrame(oat_rows, columns=oat_empty.columns)


def _scale_score_series_for_log(series: pd.Series) -> pd.Series:
    numeric_series = pd.to_numeric(series, errors="coerce")
    clean_series = numeric_series.dropna()
    if clean_series.empty:
        return numeric_series.astype(float)
    if clean_series.abs().max() <= 1.0 + 1e-9:
        return numeric_series.astype(float) * LOG_ANALYSIS_SCALE
    return numeric_series.astype(float)


def _log_transform_score_series(series: pd.Series) -> pd.Series:
    scaled_series = _scale_score_series_for_log(series).clip(lower=0)
    return np.log(scaled_series.clip(lower=LOG_ANALYSIS_EPSILON))


def _prepare_regression_frame(feature_df: pd.DataFrame, target_series: pd.Series) -> tuple[pd.DataFrame, list[str]]:
    if feature_df.empty:
        return pd.DataFrame(), []

    working_df = feature_df.copy()
    working_df["_target_"] = pd.to_numeric(target_series, errors="coerce")
    working_df = working_df.replace([np.inf, -np.inf], np.nan).dropna()
    if working_df.empty:
        return pd.DataFrame(), []

    predictor_columns = [column for column in working_df.columns if column != "_target_"]
    if not predictor_columns:
        return pd.DataFrame(), []

    design_matrix = working_df.loc[:, predictor_columns].astype(float)
    design_matrix = design_matrix.loc[:, design_matrix.nunique(dropna=True) > 0]
    if design_matrix.empty:
        return pd.DataFrame(), []

    prepared_df = pd.concat([design_matrix, working_df["_target_"]], axis=1)
    return prepared_df, design_matrix.columns.tolist()


def fit_linear_model(feature_df: pd.DataFrame, target_series: pd.Series) -> dict[str, object] | None:
    prepared_df, predictor_columns = _prepare_regression_frame(feature_df, target_series)
    if prepared_df.empty or not predictor_columns:
        return None
    if prepared_df["_target_"].nunique(dropna=True) < 2:
        return None
    if len(prepared_df) <= len(predictor_columns):
        return None

    y = prepared_df["_target_"].to_numpy(dtype=float)
    x = prepared_df.loc[:, predictor_columns].to_numpy(dtype=float)
    intercept = np.ones((x.shape[0], 1), dtype=float)
    x_with_intercept = np.hstack([intercept, x])
    coefficients, _, _, _ = np.linalg.lstsq(x_with_intercept, y, rcond=None)
    return {
        "predictor_columns": predictor_columns,
        "coefficients": coefficients,
    }


def predict_linear_model(model: dict[str, object] | None, feature_df: pd.DataFrame) -> pd.Series:
    if model is None:
        return pd.Series(np.nan, index=feature_df.index, dtype=float)

    predictor_columns = [str(column) for column in model.get("predictor_columns", [])]
    if not predictor_columns or not set(predictor_columns).issubset(feature_df.columns):
        return pd.Series(np.nan, index=feature_df.index, dtype=float)

    x = feature_df.loc[:, predictor_columns].astype(float).to_numpy(dtype=float)
    intercept = np.ones((x.shape[0], 1), dtype=float)
    x_with_intercept = np.hstack([intercept, x])
    coefficients = np.asarray(model["coefficients"], dtype=float)
    return pd.Series(x_with_intercept @ coefficients, index=feature_df.index, dtype=float)


def _inverse_log_transform_series(series: pd.Series) -> pd.Series:
    numeric_series = pd.to_numeric(series, errors="coerce")
    return np.exp(numeric_series) / LOG_ANALYSIS_SCALE


def compute_regression_r2(feature_df: pd.DataFrame, target_series: pd.Series) -> float | object:
    model = fit_linear_model(feature_df, target_series)
    if model is None:
        return pd.NA

    prepared_df, predictor_columns = _prepare_regression_frame(feature_df, target_series)
    y = prepared_df["_target_"].to_numpy(dtype=float)
    predictions = predict_linear_model(model, prepared_df.loc[:, predictor_columns]).to_numpy(dtype=float)

    total_sum_of_squares = float(((y - y.mean()) ** 2).sum())
    if np.isclose(total_sum_of_squares, 0.0):
        return pd.NA

    residual_sum_of_squares = float(((y - predictions) ** 2).sum())
    r2_value = 1.0 - (residual_sum_of_squares / total_sum_of_squares)
    return float(np.clip(r2_value, 0.0, 1.0))


def compute_exact_shapley_r2(feature_df: pd.DataFrame, target_series: pd.Series) -> tuple[dict[str, float], float | object]:
    predictor_columns = feature_df.columns.tolist()
    if not predictor_columns:
        return {}, pd.NA

    working_df = feature_df.copy()
    working_df["_target_"] = pd.to_numeric(target_series, errors="coerce")
    working_df = working_df.replace([np.inf, -np.inf], np.nan).dropna()
    if working_df.empty or working_df["_target_"].nunique(dropna=True) < 2:
        return {column: np.nan for column in predictor_columns}, pd.NA

    aligned_features = working_df.loc[:, predictor_columns].astype(float)
    aligned_target = working_df["_target_"].astype(float)
    cached_r2: dict[tuple[str, ...], float] = {tuple(): 0.0}

    def subset_r2(subset: tuple[str, ...]) -> float:
        ordered_subset = tuple(sorted(subset))
        if ordered_subset in cached_r2:
            return cached_r2[ordered_subset]
        r2_value = compute_regression_r2(aligned_features.loc[:, list(ordered_subset)], aligned_target)
        cached_r2[ordered_subset] = 0.0 if pd.isna(r2_value) else float(r2_value)
        return cached_r2[ordered_subset]

    total_r2 = subset_r2(tuple(predictor_columns))
    predictor_count = len(predictor_columns)
    shapley_map: dict[str, float] = {}

    for predictor in predictor_columns:
        other_predictors = [column for column in predictor_columns if column != predictor]
        contribution = 0.0
        for subset_size in range(len(other_predictors) + 1):
            weight = factorial(subset_size) * factorial(predictor_count - subset_size - 1) / factorial(predictor_count)
            for subset in combinations(other_predictors, subset_size):
                contribution += weight * (subset_r2(tuple(subset) + (predictor,)) - subset_r2(tuple(subset)))
        shapley_map[predictor] = float(max(contribution, 0.0))

    shapley_total = sum(shapley_map.values())
    if shapley_total > 0 and np.isfinite(total_r2):
        adjustment_ratio = float(total_r2) / shapley_total
        shapley_map = {
            predictor: float(np.clip(value * adjustment_ratio, 0.0, 1.0))
            for predictor, value in shapley_map.items()
        }

    return shapley_map, total_r2


def build_advanced_analysis_tables(
    desa_summary: pd.DataFrame,
    variable_explanation: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    empty_tables = _build_advanced_analysis_empty_tables()
    if desa_summary.empty:
        return empty_tables

    label_lookup = _resolve_analysis_label_lookup(variable_explanation)

    numeric_df = desa_summary.copy()
    core_columns = {"iid_desa", "ikd_desa"}
    for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
        core_columns.add(spec["dimension_column"])
        core_columns.update(indicator for indicator, _ in spec["indicator_specs"])
    available_columns = [column for column in core_columns if column in numeric_df.columns]
    for column in available_columns:
        numeric_df[column] = pd.to_numeric(numeric_df[column], errors="coerce")

    dimension_columns = [
        spec["dimension_column"]
        for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS
        if spec["dimension_column"] in numeric_df.columns
    ]
    if not dimension_columns:
        return empty_tables

    if "iid_desa" in numeric_df.columns and numeric_df["iid_desa"].notna().any():
        base_iid_series = pd.to_numeric(numeric_df["iid_desa"], errors="coerce")
    else:
        base_iid_series = pd.Series(0.0, index=numeric_df.index, dtype=float)
        for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
            dimension_column = spec["dimension_column"]
            if dimension_column in numeric_df.columns:
                base_iid_series = base_iid_series + (numeric_df[dimension_column].fillna(0.0) * float(spec["weight"]))
        base_iid_series = base_iid_series.clip(lower=0.0, upper=1.0)
        numeric_df["iid_desa"] = base_iid_series

    if "ikd_desa" in numeric_df.columns:
        base_ikd_series = pd.to_numeric(numeric_df["ikd_desa"], errors="coerce")
        base_ikd_series = base_ikd_series.where(base_ikd_series.notna(), 1.0 - base_iid_series)
    else:
        base_ikd_series = 1.0 - base_iid_series
        numeric_df["ikd_desa"] = base_ikd_series

    transformed_columns = {
        column: _log_transform_score_series(numeric_df[column])
        for column in available_columns + ["iid_desa", "ikd_desa"]
        if column in numeric_df.columns
    }
    transformed_df = pd.DataFrame(transformed_columns)

    dimension_rows: list[dict[str, object]] = []
    variable_rows: list[dict[str, object]] = []
    oat_rows: list[dict[str, object]] = []
    shapley_rows: list[dict[str, object]] = []

    increment_value = OAT_DIMENSION_INCREMENT

    for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
        dimension_column = spec["dimension_column"]
        if dimension_column not in transformed_df.columns:
            continue

        dimension_name = _build_dimension_display_name(dimension_column, spec["dimension_code"], label_lookup)
        dimension_rows.append(
            {
                "Dimensi": dimension_name,
                "R2 IID Desa": compute_regression_r2(
                    transformed_df.loc[:, [dimension_column]],
                    transformed_df["iid_desa"],
                ),
            }
        )

        for indicator_column, indicator_code in spec["indicator_specs"]:
            if indicator_column not in transformed_df.columns:
                continue
            variable_rows.append(
                {
                    "Dimensi": dimension_name,
                    "Variabel": _build_indicator_display_name(indicator_column, indicator_code, label_lookup),
                    "R2 Dimensi": compute_regression_r2(
                        transformed_df.loc[:, [indicator_column]],
                        transformed_df[dimension_column],
                    ),
                    "R2 IID Desa": compute_regression_r2(
                        transformed_df.loc[:, [indicator_column]],
                        transformed_df["iid_desa"],
                    ),
                }
            )

        simulated_dimensions = numeric_df.loc[:, dimension_columns].copy()
        simulated_dimensions[dimension_column] = simulated_dimensions[dimension_column].add(increment_value).clip(upper=1.0)
        simulated_iid_series = pd.Series(0.0, index=simulated_dimensions.index, dtype=float)
        for weight_spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
            weighted_dimension = weight_spec["dimension_column"]
            if weighted_dimension in simulated_dimensions.columns:
                simulated_iid_series = simulated_iid_series + (
                    simulated_dimensions[weighted_dimension].fillna(0.0) * float(weight_spec["weight"])
                )
        simulated_iid_series = simulated_iid_series.clip(lower=0.0, upper=1.0)
        base_deprivation_series = base_ikd_series.clip(lower=0.0, upper=1.0)
        simulated_deprivation_series = (1.0 - simulated_iid_series).clip(lower=0.0, upper=1.0)

        iid_delta = simulated_iid_series - base_iid_series
        deprivation_delta = simulated_deprivation_series - base_deprivation_series

        iid_increase_percent = pd.Series(np.nan, index=simulated_iid_series.index, dtype=float)
        iid_valid_mask = base_iid_series > 0
        iid_increase_percent.loc[iid_valid_mask] = (
            (simulated_iid_series.loc[iid_valid_mask] - base_iid_series.loc[iid_valid_mask])
            / base_iid_series.loc[iid_valid_mask]
        ) * 100.0
        iid_increase_percent.loc[base_iid_series.eq(0) & simulated_iid_series.eq(0)] = 0.0

        deprivation_decrease_percent = pd.Series(np.nan, index=simulated_deprivation_series.index, dtype=float)
        deprivation_valid_mask = base_deprivation_series > 0
        deprivation_decrease_percent.loc[deprivation_valid_mask] = (
            (base_deprivation_series.loc[deprivation_valid_mask] - simulated_deprivation_series.loc[deprivation_valid_mask])
            / base_deprivation_series.loc[deprivation_valid_mask]
        ) * 100.0
        deprivation_decrease_percent.loc[
            base_deprivation_series.eq(0) & simulated_deprivation_series.eq(0)
        ] = 0.0

        oat_rows.append(
            {
                "Dimensi": dimension_name,
                "Rata-rata Delta IID Desa": float(
                    iid_delta.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Rata-rata Kenaikan IID Desa (%)": float(
                    iid_increase_percent.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Rata-rata Delta Deprivasi Digital": float(
                    deprivation_delta.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Rata-rata Penurunan Deprivasi Digital (%)": float(
                    deprivation_decrease_percent.replace([np.inf, -np.inf], np.nan).dropna().mean()
                ),
                "Skenario OAT": (
                    f"+{format_numeric_label(OAT_DIMENSION_INCREMENT)} (1 persen) pada dimensi (skala 0-1), lalu dihitung perubahan IID Desa "
                    "dan deprivasi digital sebagai 1 - IID"
                ),
            }
        )

    indicator_specs_flat = [
        (spec["dimension_column"], spec["dimension_code"], indicator_column, indicator_code)
        for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS
        for indicator_column, indicator_code in spec["indicator_specs"]
        if indicator_column in transformed_df.columns
    ]
    indicator_columns_all = [indicator_column for _, _, indicator_column, _ in indicator_specs_flat]
    iid_shapley_map, iid_total_r2 = compute_exact_shapley_r2(
        transformed_df.loc[:, indicator_columns_all],
        transformed_df["iid_desa"],
    )
    iid_total_r2_float = float(iid_total_r2) if not pd.isna(iid_total_r2) else np.nan
    dimension_shapley_maps: dict[str, tuple[dict[str, float], float | object]] = {}
    for spec in ADVANCED_ANALYSIS_DIMENSION_SPECS:
        dimension_column = spec["dimension_column"]
        dimension_indicator_columns = [
            indicator_column
            for indicator_column, _ in spec["indicator_specs"]
            if indicator_column in transformed_df.columns
        ]
        if dimension_column in transformed_df.columns and dimension_indicator_columns:
            dimension_shapley_maps[dimension_column] = compute_exact_shapley_r2(
                transformed_df.loc[:, dimension_indicator_columns],
                transformed_df[dimension_column],
            )

    for dimension_column, dimension_code, indicator_column, indicator_code in indicator_specs_flat:
        dimension_name = _build_dimension_display_name(dimension_column, dimension_code, label_lookup)
        dimension_shapley_map, dimension_total_r2 = dimension_shapley_maps.get(dimension_column, ({}, pd.NA))
        dimension_total_r2_float = float(dimension_total_r2) if not pd.isna(dimension_total_r2) else np.nan
        dimension_shapley_value = float(dimension_shapley_map.get(indicator_column, np.nan))
        iid_shapley_value = float(iid_shapley_map.get(indicator_column, np.nan))
        shapley_rows.append(
            {
                "Dimensi": dimension_name,
                "Variabel": _build_indicator_display_name(indicator_column, indicator_code, label_lookup),
                "Shapley R2 Dimensi": dimension_shapley_value,
                "Proporsi Shapley Dimensi": (
                    dimension_shapley_value / dimension_total_r2_float
                    if np.isfinite(dimension_total_r2_float) and dimension_total_r2_float > 0
                    else np.nan
                ),
                "Shapley R2 IID Desa": iid_shapley_value,
                "Proporsi Shapley IID Desa": (
                    iid_shapley_value / iid_total_r2_float
                    if np.isfinite(iid_total_r2_float) and iid_total_r2_float > 0
                    else np.nan
                ),
            }
        )

    output_tables = {
        "analisis_determinasi_dimensi": pd.DataFrame(dimension_rows),
        "analisis_determinasi_variabel": pd.DataFrame(variable_rows),
        "analisis_sensitivitas_oat": pd.DataFrame(oat_rows),
        "analisis_shapley_variabel": pd.DataFrame(shapley_rows),
    }
    for key, empty_df in empty_tables.items():
        if key not in output_tables or output_tables[key].empty:
            output_tables[key] = empty_df
    return output_tables


def build_iid_rt_distribution_by_desa(
    household_master: pd.DataFrame,
    category_ranges: dict[str, str] = IID_RT_CATEGORY_RANGES,
) -> pd.DataFrame:
    group_keys = ["kode_deskel", "deskel_std"]
    valid_households = household_master.loc[household_master["valid_untuk_indeks"]].copy()
    output_columns = [
        "kode_deskel",
        "deskel",
        "kategori_iid_rt",
        "rentang_iid_rt",
        "jumlah_kk",
        "persentase_kk",
        "total_kk_desa",
    ]
    if valid_households.empty:
        return pd.DataFrame(columns=output_columns)

    desa_lookup = valid_households.groupby(group_keys, dropna=False).agg(
        deskel=("deskel", _mode_or_na),
        total_kk_desa=("family_id", "count"),
    ).reset_index()

    category_lookup = pd.DataFrame(
        {
            "kategori_iid_rt": IID_RT_CATEGORY_ORDER,
            "rentang_iid_rt": [category_ranges[label] for label in IID_RT_CATEGORY_ORDER],
            "urutan_kategori": list(range(len(IID_RT_CATEGORY_ORDER))),
        }
    )

    desa_grid = (
        desa_lookup.assign(_merge_key=1)
        .merge(category_lookup.assign(_merge_key=1), on="_merge_key", how="inner")
        .drop(columns="_merge_key")
    )

    category_counts = valid_households.groupby(group_keys + ["kategori_iid_rt"], dropna=False).agg(
        jumlah_kk=("family_id", "count")
    ).reset_index()

    distribution_df = desa_grid.merge(
        category_counts,
        on=group_keys + ["kategori_iid_rt"],
        how="left",
    )
    distribution_df["jumlah_kk"] = distribution_df["jumlah_kk"].fillna(0).astype(int)
    distribution_df["persentase_kk"] = np.where(
        distribution_df["total_kk_desa"].gt(0),
        distribution_df["jumlah_kk"] / distribution_df["total_kk_desa"],
        0.0,
    )
    distribution_df = distribution_df.sort_values(
        ["kode_deskel", "urutan_kategori"],
        kind="mergesort",
    ).reset_index(drop=True)
    distribution_df = distribution_df.drop(columns=["deskel_std", "urutan_kategori"])
    return distribution_df.loc[:, output_columns]


def build_iid_rt_distribution_summary(
    distribution_by_desa_df: pd.DataFrame,
    category_ranges: dict[str, str] = IID_RT_CATEGORY_RANGES,
) -> pd.DataFrame:
    output_columns = [
        "kategori_iid_rt",
        "rentang_iid_rt",
        "jumlah_kk",
        "persentase_kk",
        "total_kk",
    ]
    if distribution_by_desa_df.empty:
        return pd.DataFrame(columns=output_columns)

    category_lookup = pd.DataFrame(
        {
            "kategori_iid_rt": IID_RT_CATEGORY_ORDER,
            "rentang_iid_rt": [category_ranges[label] for label in IID_RT_CATEGORY_ORDER],
            "urutan_kategori": list(range(len(IID_RT_CATEGORY_ORDER))),
        }
    )
    category_counts = distribution_by_desa_df.groupby("kategori_iid_rt", dropna=False).agg(
        jumlah_kk=("jumlah_kk", "sum")
    ).reset_index()
    distribution_df = category_lookup.merge(category_counts, on="kategori_iid_rt", how="left")
    distribution_df["jumlah_kk"] = distribution_df["jumlah_kk"].fillna(0).astype(int)
    total_kk = int(distribution_df["jumlah_kk"].sum())
    distribution_df["total_kk"] = total_kk
    distribution_df["persentase_kk"] = np.where(
        total_kk > 0,
        distribution_df["jumlah_kk"] / total_kk,
        0.0,
    )
    distribution_df = distribution_df.sort_values(["urutan_kategori"], kind="mergesort").reset_index(drop=True)
    distribution_df = distribution_df.drop(columns=["urutan_kategori"])
    return distribution_df.loc[:, output_columns]


def build_iid_rt_person_distribution(
    person_df: pd.DataFrame,
    household_master: pd.DataFrame,
    category_ranges: dict[str, str] = IID_RT_CATEGORY_RANGES,
) -> pd.DataFrame:
    output_columns = [
        "kategori_iid_rt",
        "rentang_iid_rt",
        "jumlah_warga",
        "persentase_warga",
        "total_warga",
    ]
    if person_df.empty:
        return pd.DataFrame(columns=output_columns)

    warga_df = person_df.copy()
    warga_df["family_id"] = derive_family_id(warga_df)
    warga_df = warga_df.merge(
        household_master[["family_id", "valid_untuk_indeks", "kategori_iid_rt"]],
        on="family_id",
        how="left",
    )
    warga_df["kategori_iid_rt"] = warga_df["kategori_iid_rt"].where(
        warga_df["valid_untuk_indeks"].eq(True),
        UNSCORED_IID_CATEGORY_LABEL,
    )
    order = IID_RT_CATEGORY_ORDER + [UNSCORED_IID_CATEGORY_LABEL]
    counts = (
        warga_df["kategori_iid_rt"]
        .value_counts(dropna=False)
        .reindex(order, fill_value=0)
        .rename("jumlah_warga")
        .reset_index()
        .rename(columns={"index": "kategori_iid_rt"})
    )
    total_warga = int(counts["jumlah_warga"].sum())
    counts["rentang_iid_rt"] = counts["kategori_iid_rt"].map(category_ranges).fillna("RT tidak valid/tidak diberi skor")
    counts["persentase_warga"] = np.where(total_warga > 0, counts["jumlah_warga"] / total_warga, 0.0)
    counts["total_warga"] = total_warga
    return counts.loc[:, output_columns]


def build_small_count_distribution(
    series: pd.Series,
    max_explicit: int,
    overflow_label: str,
    count_label: str,
) -> pd.DataFrame:
    clean = pd.to_numeric(series, errors="coerce").fillna(0).clip(lower=0)
    bucket_order = [str(i) for i in range(max_explicit + 1)] + [overflow_label]
    bucketed = clean.apply(lambda value: overflow_label if int(value) > max_explicit else str(int(value)))
    counts = bucketed.value_counts().reindex(bucket_order, fill_value=0)
    total = int(counts.sum())
    return pd.DataFrame(
        {
            "kategori": bucket_order,
            count_label: counts.astype(int).values,
            "persentase_rt": np.where(total > 0, counts.values / total, 0.0),
            "total_rt": total,
        }
    )


def build_spend_distribution(series: pd.Series, count_label: str) -> pd.DataFrame:
    clean = pd.to_numeric(series, errors="coerce").fillna(0).clip(lower=0)
    bins = [
        ("0", lambda value: value <= 0),
        ("1-49.999", lambda value: 0 < value < 50000),
        ("50.000-99.999", lambda value: 50000 <= value < 100000),
        ("100.000-199.999", lambda value: 100000 <= value < 200000),
        ("200.000-499.999", lambda value: 200000 <= value < 500000),
        (">=500.000", lambda value: value >= 500000),
    ]
    labels = []
    for value in clean:
        assigned = "lainnya"
        for label, matcher in bins:
            if matcher(float(value)):
                assigned = label
                break
        labels.append(assigned)
    order = [label for label, _ in bins]
    counts = pd.Series(labels, dtype="string").value_counts().reindex(order, fill_value=0)
    total = int(counts.sum())
    return pd.DataFrame(
        {
            "kategori": order,
            count_label: counts.astype(int).values,
            "persentase_rt": np.where(total > 0, counts.values / total, 0.0),
            "total_rt": total,
        }
    )


def build_tinggi_profile_tables(household_master: pd.DataFrame) -> dict[str, pd.DataFrame]:
    valid_high = household_master.loc[
        household_master["valid_untuk_indeks"].fillna(False) & household_master["kategori_iid_rt"].eq("tinggi")
    ].copy()
    if valid_high.empty:
        empty_summary = pd.DataFrame(columns=["metrik", "nilai"])
        empty_dist = pd.DataFrame(columns=["kategori", "jumlah_rt", "persentase_rt", "total_rt"])
        return {
            "summary": empty_summary,
            "hp": empty_dist,
            "provider": empty_dist,
            "device": empty_dist,
            "spend": empty_dist,
        }

    hp_count = pd.to_numeric(valid_high["hp_jumlah_terstandar"], errors="coerce").fillna(0)
    provider_count = valid_high["hp_provider_teragregasi"].apply(count_multivalue_items).astype(int)
    device_count = pd.to_numeric(valid_high["jumlah_perangkat_produktif_rumah_tangga"], errors="coerce").fillna(0)
    spend_value = pd.to_numeric(valid_high["rp_komunikasi_tertinggi"], errors="coerce").fillna(0)
    iid_value = pd.to_numeric(valid_high["iid_rumah_tangga"], errors="coerce")

    summary_df = pd.DataFrame(
        [
            {"metrik": "Jumlah RT kategori tinggi", "nilai": int(len(valid_high))},
            {"metrik": "Rata-rata IID-RT", "nilai": float(iid_value.mean())},
            {"metrik": "Rata-rata jumlah HP", "nilai": float(hp_count.mean())},
            {"metrik": "Median jumlah HP", "nilai": float(hp_count.median())},
            {"metrik": "Rata-rata jumlah provider HP", "nilai": float(provider_count.mean())},
            {"metrik": "Median jumlah provider HP", "nilai": float(provider_count.median())},
            {"metrik": "Rata-rata jumlah laptop/perangkat produktif", "nilai": float(device_count.mean())},
            {"metrik": "Median jumlah laptop/perangkat produktif", "nilai": float(device_count.median())},
            {"metrik": "Rata-rata biaya komunikasi", "nilai": float(spend_value.mean())},
            {"metrik": "Median biaya komunikasi", "nilai": float(spend_value.median())},
            {"metrik": "Total biaya komunikasi", "nilai": float(spend_value.sum())},
        ]
    )

    return {
        "summary": summary_df,
        "hp": build_small_count_distribution(hp_count, max_explicit=5, overflow_label="6+", count_label="jumlah_rt"),
        "provider": build_small_count_distribution(
            provider_count, max_explicit=3, overflow_label="4+", count_label="jumlah_rt"
        ),
        "device": build_small_count_distribution(
            device_count, max_explicit=3, overflow_label="4+", count_label="jumlah_rt"
        ),
        "spend": build_spend_distribution(spend_value, count_label="jumlah_rt"),
    }


def build_gini_distribution_summary(desa_summary: pd.DataFrame) -> pd.DataFrame:
    _, distribution_df = apply_relative_gini_classification(desa_summary)
    return distribution_df


def add_gini_visualization_to_indeks_desa(workbook: object, gini_distribution_df: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title="visualisasi_gini")
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.zoomScale = 90

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 20
    worksheet.column_dimensions["H"].width = 16

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    section_font = Font(bold=True, size=11)
    title_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    header_fill = PatternFill(fill_type="solid", fgColor="FBE5D6")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    title_cell = worksheet.cell(row=1, column=1, value="Visualisasi Sebaran Gini IID-RT Antar Desa")
    title_cell.font = title_font
    title_cell.fill = title_fill

    worksheet["A2"] = (
        "Pie chart menunjukkan proporsi desa pada tiap kategori relatif ketimpangan, sedangkan diagram batang menunjukkan jumlah desanya."
    )
    worksheet["A3"] = (
        "Karena seluruh Gini desa berada pada rentang rendah secara absolut, interpretasi memakai tertil relatif antar desa dalam sampel: sepertiga desa dengan Gini terendah = Rendah, sepertiga berikutnya = Sedang, dan sepertiga tertinggi = Tinggi."
    )
    worksheet["A2"].alignment = wrap_alignment
    worksheet["A3"].alignment = wrap_alignment

    if gini_distribution_df.empty:
        worksheet["A5"] = "Tidak ada data desa untuk divisualisasikan."
        return

    total_desa = int(gini_distribution_df["total_desa"].iloc[0])
    dominant_row = gini_distribution_df.sort_values(
        ["jumlah_desa", "persentase_desa", "interpretasi_gini"],
        ascending=[False, False, True],
        kind="mergesort",
    ).iloc[0]

    worksheet.cell(row=4, column=1, value="Total desa").font = header_font
    worksheet.cell(row=4, column=2, value=total_desa)
    worksheet.cell(row=4, column=3, value="Kategori dominan").font = header_font
    worksheet.cell(
        row=4,
        column=4,
        value=f"{dominant_row['interpretasi_gini']} ({dominant_row['rentang_gini']})",
    ).alignment = wrap_alignment

    headers = ["Kategori Relatif", "Rentang Gini", "Jumlah Desa", "% dari total desa"]
    header_row = 5
    for offset, header in enumerate(headers):
        header_cell = worksheet.cell(row=header_row, column=1 + offset, value=header)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = wrap_alignment

    for row_offset, row in enumerate(gini_distribution_df.itertuples(index=False), start=1):
        data_row = header_row + row_offset
        worksheet.cell(row=data_row, column=1, value=row.interpretasi_gini)
        worksheet.cell(row=data_row, column=2, value=row.rentang_gini)
        worksheet.cell(row=data_row, column=3, value=int(row.jumlah_desa))
        percentage_cell = worksheet.cell(row=data_row, column=4, value=float(row.persentase_desa))
        percentage_cell.number_format = EXCEL_PERCENT_FORMAT
        percentage_cell.alignment = wrap_alignment

    data_reference = Reference(
        worksheet,
        min_col=3,
        min_row=header_row,
        max_row=header_row + len(gini_distribution_df),
    )
    label_reference = Reference(
        worksheet,
        min_col=1,
        min_row=header_row + 1,
        max_row=header_row + len(gini_distribution_df),
    )

    pie_chart = PieChart()
    pie_chart.title = "Sebaran Kategori Relatif Gini Antar Desa"
    pie_chart.height = 8
    pie_chart.width = 10
    pie_chart.varyColors = True
    pie_chart.legend.position = "r"
    pie_chart.add_data(data_reference, titles_from_data=True)
    pie_chart.set_categories(label_reference)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.dataLabels.showCatName = True
    worksheet.add_chart(pie_chart, "F4")

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Jumlah Desa per Kategori Relatif Gini"
    bar_chart.y_axis.title = "Jumlah Desa"
    bar_chart.x_axis.title = "Kategori Ketimpangan"
    bar_chart.height = 8
    bar_chart.width = 12
    bar_chart.varyColors = True
    bar_chart.legend = None
    bar_chart.dataLabels = DataLabelList()
    bar_chart.dataLabels.showVal = True
    bar_chart.add_data(data_reference, titles_from_data=True)
    bar_chart.set_categories(label_reference)
    worksheet.add_chart(bar_chart, "N4")

    detail_start_row = header_row + len(gini_distribution_df) + 4
    worksheet.merge_cells(start_row=detail_start_row, start_column=1, end_row=detail_start_row, end_column=4)
    detail_title_cell = worksheet.cell(row=detail_start_row, column=1, value="Cara Baca Distribusi Gini")
    detail_title_cell.font = section_font
    detail_title_cell.fill = title_fill
    worksheet.cell(
        row=detail_start_row + 1,
        column=1,
        value=(
            "Klasifikasi ini bersifat relatif di dalam sampel penelitian. Jadi label Rendah, Sedang, dan Tinggi menunjukkan posisi ketimpangan antar desa, "
            "bukan ambang absolut seperti 0,30 atau 0,50."
        ),
    ).alignment = wrap_alignment


def add_iid_rt_visualization_sheet(
    workbook: object,
    person_distribution_df: pd.DataFrame,
    distribution_by_desa_df: pd.DataFrame,
) -> None:
    worksheet = workbook.create_sheet(title="visualisasi_iid_rt")
    worksheet.freeze_panes = "A6"
    worksheet.sheet_view.zoomScale = 90

    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 4
    worksheet.column_dimensions["H"].width = 18
    worksheet.column_dimensions["I"].width = 18
    worksheet.column_dimensions["J"].width = 18

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    header_font = Font(bold=True)
    title_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_fill = PatternFill(fill_type="solid", fgColor="EAF3E2")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    worksheet["A1"] = "Visualisasi IID-RT"
    worksheet["A1"].font = title_font
    worksheet["A2"] = "Diagram batang menunjukkan sebaran seluruh warga pada kategori IID-RT rumah tangga masing-masing."
    worksheet["A3"] = "Pie chart dibuat untuk setiap desa agar komposisi kategori IID-RT per desa bisa dibaca cepat."
    worksheet["A4"] = (
        "Rentang kategori: sangat rendah 0,00-0,20 | rendah >0,20-0,40 | sedang >0,40-0,60 | "
        "tinggi >0,60-0,80 | sangat tinggi >0,80-1,00"
    )
    worksheet["A2"].alignment = wrap_alignment
    worksheet["A3"].alignment = wrap_alignment
    worksheet["A4"].alignment = wrap_alignment

    if person_distribution_df.empty and distribution_by_desa_df.empty:
        worksheet["A6"] = "Tidak ada data IID-RT yang bisa divisualisasikan."
        return

    summary_header_row = 6
    if person_distribution_df.empty:
        worksheet["A6"] = "Tidak ada data warga yang bisa diringkas pada diagram batang."
        detail_start_row = 12
    else:
        total_warga = int(person_distribution_df["total_warga"].iloc[0])
        dominant_row = person_distribution_df.sort_values(
        ["jumlah_warga", "persentase_warga", "kategori_iid_rt"],
        ascending=[False, False, True],
        kind="mergesort",
        ).iloc[0]

        worksheet.cell(row=5, column=1, value="Total warga").font = header_font
        worksheet.cell(row=5, column=2, value=total_warga)
        worksheet.cell(row=5, column=3, value="Kategori dominan").font = header_font
        worksheet.cell(
            row=5,
            column=4,
            value=f"{dominant_row['kategori_iid_rt']} ({dominant_row['rentang_iid_rt']})",
        ).alignment = wrap_alignment

        summary_headers = ["Kategori", "Rentang IID-RT", "Jumlah Warga", "% dari total warga"]
        for column_index, header in enumerate(summary_headers, start=1):
            header_cell = worksheet.cell(row=summary_header_row, column=column_index, value=header)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = wrap_alignment

        for row_offset, row in enumerate(person_distribution_df.itertuples(index=False), start=1):
            data_row = summary_header_row + row_offset
            worksheet.cell(row=data_row, column=1, value=row.kategori_iid_rt)
            worksheet.cell(row=data_row, column=2, value=row.rentang_iid_rt)
            worksheet.cell(row=data_row, column=3, value=int(row.jumlah_warga))
            percentage_cell = worksheet.cell(row=data_row, column=4, value=float(row.persentase_warga))
            percentage_cell.number_format = EXCEL_PERCENT_FORMAT
            percentage_cell.alignment = wrap_alignment

        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Sebaran Seluruh Warga"
        bar_chart.y_axis.title = "Jumlah Warga"
        bar_chart.x_axis.title = "Kategori IID-RT"
        bar_chart.height = 8
        bar_chart.width = 12
        bar_chart.varyColors = True
        bar_chart.legend = None
        bar_chart.dataLabels = DataLabelList()
        bar_chart.dataLabels.showVal = True

        data_reference = Reference(
            worksheet,
            min_col=3,
            min_row=summary_header_row,
            max_row=summary_header_row + len(person_distribution_df),
        )
        label_reference = Reference(
            worksheet,
            min_col=1,
            min_row=summary_header_row + 1,
            max_row=summary_header_row + len(person_distribution_df),
        )
        bar_chart.add_data(data_reference, titles_from_data=True)
        bar_chart.set_categories(label_reference)
        worksheet.add_chart(bar_chart, "F4")
        detail_start_row = summary_header_row + len(person_distribution_df) + 5

    worksheet.merge_cells(start_row=detail_start_row, start_column=1, end_row=detail_start_row, end_column=6)
    detail_title_cell = worksheet.cell(row=detail_start_row, column=1, value="Rincian Sebaran IID-RT per Desa")
    detail_title_cell.font = section_font
    detail_title_cell.fill = title_fill

    detail_headers = ["Kode Deskel", "Desa", "Kategori", "Rentang IID-RT", "Jumlah KK", "% dari total KK desa"]
    detail_header_row = detail_start_row + 1
    for column_index, header in enumerate(detail_headers, start=1):
        header_cell = worksheet.cell(row=detail_header_row, column=column_index, value=header)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = wrap_alignment

    if distribution_by_desa_df.empty:
        worksheet.cell(row=detail_header_row + 1, column=1, value="Tidak ada data desa untuk divisualisasikan.")
        return

    for row_offset, row in enumerate(distribution_by_desa_df.itertuples(index=False), start=1):
        data_row = detail_header_row + row_offset
        worksheet.cell(row=data_row, column=1, value=row.kode_deskel)
        worksheet.cell(row=data_row, column=2, value=row.deskel)
        worksheet.cell(row=data_row, column=3, value=row.kategori_iid_rt)
        worksheet.cell(row=data_row, column=4, value=row.rentang_iid_rt)
        worksheet.cell(row=data_row, column=5, value=int(row.jumlah_kk))
        percentage_cell = worksheet.cell(row=data_row, column=6, value=float(row.persentase_kk))
        percentage_cell.number_format = EXCEL_PERCENT_FORMAT
        percentage_cell.alignment = wrap_alignment

    pie_title_row = detail_start_row
    worksheet.merge_cells(start_row=pie_title_row, start_column=8, end_row=pie_title_row, end_column=18)
    pie_title_cell = worksheet.cell(row=pie_title_row, column=8, value="Pie Chart Kategori IID-RT per Desa")
    pie_title_cell.font = section_font
    pie_title_cell.fill = title_fill

    current_data_row = detail_header_row + 1
    for chart_index, (desa_key, desa_block) in enumerate(
        distribution_by_desa_df.groupby(["kode_deskel", "deskel"], dropna=False, sort=False)
    ):
        _, desa_name = desa_key
        data_start_row = current_data_row
        data_end_row = current_data_row + len(desa_block) - 1
        chart = PieChart()
        chart.title = str(desa_name)[:40] if pd.notna(desa_name) else "Desa"
        chart.height = 6
        chart.width = 7
        chart.varyColors = True
        chart.legend.position = "r"
        chart_data_reference = Reference(
            worksheet,
            min_col=5,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart_label_reference = Reference(
            worksheet,
            min_col=3,
            min_row=data_start_row,
            max_row=data_end_row,
        )
        chart.add_data(chart_data_reference, titles_from_data=False)
        chart.set_categories(chart_label_reference)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True

        chart_column = 8 + (chart_index % 3) * 7
        chart_row = pie_title_row + 2 + (chart_index // 3) * 16
        worksheet.add_chart(chart, f"{get_column_letter(chart_column)}{chart_row}")
        current_data_row = data_end_row + 1


def add_tinggi_profile_sheet(workbook: object, profile_tables: dict[str, pd.DataFrame]) -> None:
    worksheet = workbook.create_sheet(title="profil_iid_tinggi")
    worksheet.freeze_panes = "A5"
    worksheet.sheet_view.zoomScale = 90

    worksheet.column_dimensions["A"].width = 36
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 16
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 14
    worksheet.column_dimensions["H"].width = 14

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    header_font = Font(bold=True)
    title_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    header_fill = PatternFill(fill_type="solid", fgColor="F3F9EC")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    summary_df = profile_tables["summary"]
    hp_df = profile_tables["hp"]
    provider_df = profile_tables["provider"]
    device_df = profile_tables["device"]
    spend_df = profile_tables["spend"]

    worksheet["A1"] = "Profil Rumah Tangga Kategori Tinggi"
    worksheet["A1"].font = title_font
    worksheet["A2"] = (
        "Sheet ini merangkum rumah tangga dengan kategori IID-RT = tinggi, termasuk jumlah HP, jumlah provider HP, "
        "jumlah laptop/perangkat produktif, dan biaya komunikasi."
    )
    worksheet["A2"].alignment = wrap_alignment

    if summary_df.empty:
        worksheet["A4"] = "Tidak ada rumah tangga dengan kategori tinggi."
        return

    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    summary_title = worksheet.cell(row=4, column=1, value="Ringkasan Utama")
    summary_title.font = section_font
    summary_title.fill = title_fill

    summary_header_row = 5
    for column_index, header in enumerate(["Metrik", "Nilai"], start=1):
        header_cell = worksheet.cell(row=summary_header_row, column=column_index, value=header)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = wrap_alignment

    for row_offset, row in enumerate(summary_df.itertuples(index=False), start=1):
        data_row = summary_header_row + row_offset
        worksheet.cell(row=data_row, column=1, value=row.metrik)
        value_cell = worksheet.cell(row=data_row, column=2, value=row.nilai)
        if isinstance(row.nilai, Number):
            numeric_value = float(row.nilai)
            if np.isfinite(numeric_value) and np.isclose(numeric_value, round(numeric_value)):
                value_cell.value = int(round(numeric_value))
                value_cell.number_format = "0"
            else:
                value_cell.number_format = EXCEL_FLOAT_FORMAT

    def add_distribution_section(
        start_row: int,
        title: str,
        label_header: str,
        table_df: pd.DataFrame,
        chart_title: str,
        value_axis_title: str,
    ) -> None:
        worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        title_cell = worksheet.cell(row=start_row, column=1, value=title)
        title_cell.font = section_font
        title_cell.fill = title_fill

        header_row = start_row + 1
        headers = [label_header, "Jumlah RT", "% RT"]
        for column_index, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=header_row, column=column_index, value=header)
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = wrap_alignment

        for row_offset, row in enumerate(table_df.itertuples(index=False), start=1):
            data_row = header_row + row_offset
            worksheet.cell(row=data_row, column=1, value=row.kategori)
            worksheet.cell(row=data_row, column=2, value=int(row.jumlah_rt))
            pct_cell = worksheet.cell(row=data_row, column=3, value=float(row.persentase_rt))
            pct_cell.number_format = EXCEL_PERCENT_FORMAT

        data_reference = Reference(
            worksheet,
            min_col=2,
            min_row=header_row,
            max_row=header_row + len(table_df),
        )
        label_reference = Reference(
            worksheet,
            min_col=1,
            min_row=header_row + 1,
            max_row=header_row + len(table_df),
        )
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_title
        chart.y_axis.title = value_axis_title
        chart.x_axis.title = label_header
        chart.height = 6
        chart.width = 10
        chart.varyColors = True
        chart.legend = None
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.add_data(data_reference, titles_from_data=True)
        chart.set_categories(label_reference)
        worksheet.add_chart(chart, f"F{start_row}")

    add_distribution_section(18, "Distribusi Jumlah HP", "Jumlah HP", hp_df, "Distribusi Jumlah HP", "Jumlah RT")
    add_distribution_section(
        32,
        "Distribusi Jumlah Provider HP",
        "Jumlah Provider",
        provider_df,
        "Distribusi Jumlah Provider HP",
        "Jumlah RT",
    )
    add_distribution_section(
        46,
        "Distribusi Jumlah Laptop/Perangkat Produktif",
        "Jumlah Perangkat",
        device_df,
        "Distribusi Laptop/Perangkat Produktif",
        "Jumlah RT",
    )
    add_distribution_section(
        60,
        "Distribusi Biaya Komunikasi",
        "Rentang Biaya",
        spend_df,
        "Distribusi Biaya Komunikasi",
        "Jumlah RT",
    )


def build_variable_explanation(
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> pd.DataFrame:
    rows = [
        {
            "nama_variabel": "indikator_A",
            "level_output": "keluarga, desa",
            "label_konsep": "Kepemilikan HP rumah tangga",
            "dimensi": "Akses perangkat",
            "simbol_dimensi": "A_h",
            "bobot_dimensi": 0.25,
            "sumber_nilai": "hp_punya, hp_jumlah",
            "aturan_skoring": "0 = tidak memiliki HP; 1 = memiliki minimal satu HP",
        },
        {
            "nama_variabel": "indikator_B",
            "level_output": "keluarga, desa",
            "label_konsep": "Kecukupan HP",
            "dimensi": "Akses perangkat",
            "simbol_dimensi": "A_h",
            "bobot_dimensi": 0.25,
            "sumber_nilai": "hp_jumlah, jml_keluarga",
            "aturan_skoring": "min(hp_jumlah / jumlah anggota keluarga, 1)",
        },
        {
            "nama_variabel": "indikator_C",
            "level_output": "keluarga, desa",
            "label_konsep": "Kepemilikan perangkat digital produktif",
            "dimensi": "Akses perangkat",
            "simbol_dimensi": "A_h",
            "bobot_dimensi": 0.25,
            "sumber_nilai": "elektronik_rumah",
            "aturan_skoring": "0 = tidak memiliki laptop/notebook/komputer; 1 = memiliki minimal satu perangkat digital produktif",
        },
        {
            "nama_variabel": "indikator_D",
            "level_output": "keluarga, desa",
            "label_konsep": "Akses internet rumah tangga",
            "dimensi": "Konektivitas internet",
            "simbol_dimensi": "K_h",
            "bobot_dimensi": 0.25,
            "sumber_nilai": "wifi, hp_provider, rp_komunikasi",
            "aturan_skoring": (
                "0 = tidak ada akses; 0,5 = akses publik/gratis/terbatas; 0,75 = paket data seluler/provider HP; "
                "1 = Wi-Fi/langganan internet rumah tangga"
            ),
        },
        {
            "nama_variabel": "indikator_E",
            "level_output": "keluarga, desa",
            "label_konsep": "Pendidikan terakhir kepala keluarga",
            "dimensi": "Kapasitas manusia",
            "simbol_dimensi": "M_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "status_dalam_keluarga, ijazah",
            "aturan_skoring": (
                "Tidak punya ijazah = 0; TK/PAUD = 0,10; SD = 0,25; SMP = 0,50; SMA/SMK = 0,75; "
                "D1/D2/D3 = 0,85; D4/S1 = 0,95; S2/S3 = 1"
            ),
        },
        {
            "nama_variabel": "indikator_F",
            "level_output": "keluarga, desa",
            "label_konsep": "Rasio partisipasi sekolah",
            "dimensi": "Kapasitas manusia",
            "simbol_dimensi": "M_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "usia, partisipasi_sekolah",
            "aturan_skoring": (
                f"jumlah anggota {build_school_age_range_text(school_age_min, school_age_max)} yang sedang sekolah "
                f"/ jumlah anggota {build_school_age_range_text(school_age_min, school_age_max)}; "
                "jika tidak ada anggota pada rentang usia tersebut maka NA"
            ),
        },
        {
            "nama_variabel": "indikator_G",
            "level_output": "keluarga, desa",
            "label_konsep": "Keterlibatan organisasi kepala keluarga",
            "dimensi": "Lingkungan pendukung sosial",
            "simbol_dimensi": "L_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "kepala keluarga",
            "aturan_skoring": "0 = 0; 1 organisasi = 0,5; >1 organisasi = 1",
        },
        {
            "nama_variabel": "indikator_H",
            "level_output": "keluarga, desa",
            "label_konsep": "Keterlibatan organisasi anggota keluarga",
            "dimensi": "Lingkungan pendukung sosial",
            "simbol_dimensi": "L_h",
            "bobot_dimensi": 0.10,
            "sumber_nilai": "par_organisasi, organisasi_nama",
            "aturan_skoring": "0 = tidak ada anggota ikut organisasi; 0,5 = ada satu anggota/satu organisasi; 1 = lebih dari satu anggota/lebih dari satu organisasi",
        },
        {
            "nama_variabel": "indikator_I",
            "level_output": "keluarga, desa",
            "label_konsep": "Partisipasi kepala keluarga pada kegiatan masyarakat",
            "dimensi": "Lingkungan pendukung sosial",
            "simbol_dimensi": "L_h",
            "bobot_dimensi": 0.10,
            "sumber_nilai": "status_dalam_keluarga, par_masyarakat",
            "aturan_skoring": "0 = tidak berpartisipasi; 0,5 = satu kegiatan/terbatas; 1 = lebih dari satu kegiatan/aktif",
        },
        {
            "nama_variabel": "indikator_J",
            "level_output": "keluarga, desa",
            "label_konsep": "Partisipasi anggota keluarga pada kegiatan masyarakat",
            "dimensi": "Lingkungan pendukung sosial",
            "simbol_dimensi": "L_h",
            "bobot_dimensi": 0.10,
            "sumber_nilai": "par_masyarakat",
            "aturan_skoring": "0 = tidak ada anggota berpartisipasi; 0,5 = satu anggota/satu kegiatan; 1 = lebih dari satu anggota/lebih dari satu kegiatan",
        },
        {
            "nama_variabel": "indikator_K",
            "level_output": "keluarga, desa",
            "label_konsep": "Penggunaan media sosial",
            "dimensi": "Penggunaan digital",
            "simbol_dimensi": "P_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "medsos",
            "aturan_skoring": "0 = tidak menggunakan; 0,5 = penggunaan terbatas; 1 = aktif menggunakan media sosial",
        },
        {
            "nama_variabel": "indikator_L",
            "level_output": "keluarga, desa",
            "label_konsep": "Akses media informasi",
            "dimensi": "Penggunaan digital",
            "simbol_dimensi": "P_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "media_informasi",
            "aturan_skoring": "0 = tidak mengakses media informasi; 0,5 = hanya media non-digital; 1 = mengakses media digital/online",
        },
        {
            "nama_variabel": "indikator_M",
            "level_output": "keluarga, desa",
            "label_konsep": "Partisipasi informasi/kebijakan",
            "dimensi": "Penggunaan digital",
            "simbol_dimensi": "P_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "par_kebijakan",
            "aturan_skoring": "0 = tidak pernah terlibat; 0,5 = pernah/terbatas/satu orang; 1 = aktif atau lebih dari satu orang",
        },
        {
            "nama_variabel": "dimensi_A",
            "level_output": "keluarga, desa",
            "label_konsep": "Akses perangkat",
            "dimensi": "Akses perangkat",
            "simbol_dimensi": "A_h",
            "bobot_dimensi": 0.25,
            "sumber_nilai": "rerata indikator_A, indikator_B, indikator_C",
            "aturan_skoring": "mean(indikator_A, indikator_B, indikator_C)",
        },
        {
            "nama_variabel": "dimensi_B",
            "level_output": "keluarga, desa",
            "label_konsep": "Konektivitas internet",
            "dimensi": "Konektivitas internet",
            "simbol_dimensi": "K_h",
            "bobot_dimensi": 0.25,
            "sumber_nilai": "indikator_D",
            "aturan_skoring": "indikator_D",
        },
        {
            "nama_variabel": "dimensi_C",
            "level_output": "keluarga, desa",
            "label_konsep": "Kapasitas manusia",
            "dimensi": "Kapasitas manusia",
            "simbol_dimensi": "M_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "rerata indikator_E, indikator_F",
            "aturan_skoring": "mean(indikator_E, indikator_F); indikator_F diabaikan dari penyebut jika NA",
        },
        {
            "nama_variabel": "dimensi_D",
            "level_output": "keluarga, desa",
            "label_konsep": "Penggunaan digital",
            "dimensi": "Penggunaan digital",
            "simbol_dimensi": "P_h",
            "bobot_dimensi": 0.20,
            "sumber_nilai": "rerata indikator_K, indikator_L, indikator_M",
            "aturan_skoring": "mean(indikator_K, indikator_L, indikator_M)",
        },
        {
            "nama_variabel": "dimensi_E",
            "level_output": "keluarga, desa",
            "label_konsep": "Lingkungan pendukung sosial",
            "dimensi": "Lingkungan pendukung sosial",
            "simbol_dimensi": "L_h",
            "bobot_dimensi": 0.10,
            "sumber_nilai": "rerata indikator_G, indikator_H, indikator_I, indikator_J",
            "aturan_skoring": "mean(indikator_G, indikator_H, indikator_I, indikator_J)",
        },
        {
            "nama_variabel": "iid_rumah_tangga",
            "level_output": "keluarga",
            "label_konsep": "Indeks Inklusi Digital rumah tangga",
            "dimensi": "gabungan seluruh dimensi",
            "simbol_dimensi": "IID-RT",
            "bobot_dimensi": pd.NA,
            "sumber_nilai": "gabungan dimensi_A sampai dimensi_E",
            "aturan_skoring": "0,25*dimensi_A + 0,25*dimensi_B + 0,20*dimensi_C + 0,20*dimensi_D + 0,10*dimensi_E",
        },
        {
            "nama_variabel": "kategori_iid_rt",
            "level_output": "keluarga",
            "label_konsep": "Kategori IID-RT",
            "dimensi": "klasifikasi rumah tangga",
            "simbol_dimensi": "Kategori IID-RT",
            "bobot_dimensi": pd.NA,
            "sumber_nilai": "turunan dari iid_rumah_tangga",
            "aturan_skoring": (
                f"sangat rendah {IID_RT_CATEGORY_RANGES['sangat rendah']}; "
                f"rendah {IID_RT_CATEGORY_RANGES['rendah']}; "
                f"sedang {IID_RT_CATEGORY_RANGES['sedang']}; "
                f"tinggi {IID_RT_CATEGORY_RANGES['tinggi']}; "
                f"sangat tinggi {IID_RT_CATEGORY_RANGES['sangat tinggi']}"
            ),
        },
        {
            "nama_variabel": "iid_desa",
            "level_output": "desa",
            "label_konsep": "Indeks Inklusi Digital desa",
            "dimensi": "agregasi rumah tangga",
            "simbol_dimensi": "IID-Desa",
            "bobot_dimensi": pd.NA,
            "sumber_nilai": "agregasi seluruh KK valid di desa",
            "aturan_skoring": "rerata iid_rumah_tangga pada setiap desa",
        },
        {
            "nama_variabel": "ikd_desa",
            "level_output": "desa",
            "label_konsep": "Indeks deprivasi/kesenjangan digital desa",
            "dimensi": "agregasi rumah tangga",
            "simbol_dimensi": "IKD-Desa",
            "bobot_dimensi": pd.NA,
            "sumber_nilai": "turunan dari iid_desa",
            "aturan_skoring": "1 - iid_desa; bukan Indeks Kesejahteraan Desa, melainkan komplemen untuk membaca deprivasi digital",
        },
        {
            "nama_variabel": "gini_iid_rumah_tangga",
            "level_output": "desa",
            "label_konsep": "Ketimpangan internal skor rumah tangga dalam desa",
            "dimensi": "agregasi rumah tangga",
            "simbol_dimensi": "Gini IID-RT",
            "bobot_dimensi": pd.NA,
            "sumber_nilai": "distribusi iid_rumah_tangga dalam setiap desa",
            "aturan_skoring": "koefisien Gini atas iid_rumah_tangga pada setiap desa",
        },
        {
            "nama_variabel": "interpretasi_gini",
            "level_output": "desa",
            "label_konsep": "Kategori relatif Gini IID-RT",
            "dimensi": "klasifikasi relatif ketimpangan desa",
            "simbol_dimensi": "Kategori relatif Gini",
            "bobot_dimensi": pd.NA,
            "sumber_nilai": "turunan dari gini_iid_rumah_tangga",
            "aturan_skoring": GINI_INTERPRETATION_RULE_TEXT,
        },
    ]
    explanation_df = pd.DataFrame(rows)
    explanation_df["catatan"] = pd.NA
    explanation_df.loc[explanation_df["nama_variabel"] == "indikator_F", "catatan"] = (
        "Jika tidak ada anggota usia sekolah maka indikator bernilai NA dan tidak dimasukkan ke penyebut dimensi."
    )
    explanation_df.loc[explanation_df["nama_variabel"] == "iid_rumah_tangga", "catatan"] = (
        "Skor akhir dibentuk dari rerata indikator valid pada setiap dimensi, lalu lima dimensi dibobotkan 0,25; 0,25; 0,20; 0,20; 0,10."
    )
    explanation_df["tampil_pada_baris"] = explanation_df["nama_variabel"].apply(
        lambda value: (
            build_school_indicator_row_text(school_age_min, school_age_max)
            if value == "indikator_F"
            else "kepala keluarga saja"
            if str(value).startswith("indikator_")
            else "kepala keluarga saja dalam data_keluarga"
            if value in {"dimensi_A", "dimensi_B", "dimensi_C", "dimensi_D", "dimensi_E", "iid_rumah_tangga", "kategori_iid_rt"}
            else "agregat desa"
        )
    )
    return explanation_df


def apply_recommended_scheme(
    valid_households: pd.DataFrame,
) -> tuple[pd.DataFrame, tuple[float, float, float, float], dict[str, str]]:
    recommended_df = valid_households.copy()
    recommended_df["dimensi_akses_perangkat"] = compute_weighted_dimension(
        recommended_df,
        RECOMMENDED_INDICATOR_WEIGHTS["dimensi_akses_perangkat"],
    )
    recommended_df["dimensi_konektivitas"] = compute_weighted_dimension(
        recommended_df,
        RECOMMENDED_INDICATOR_WEIGHTS["dimensi_konektivitas"],
    )
    recommended_df["dimensi_kapasitas_manusia"] = compute_weighted_dimension(
        recommended_df,
        RECOMMENDED_INDICATOR_WEIGHTS["dimensi_kapasitas_manusia"],
    )
    recommended_df["dimensi_penggunaan_digital"] = compute_weighted_dimension(
        recommended_df,
        RECOMMENDED_INDICATOR_WEIGHTS["dimensi_penggunaan_digital"],
    )
    recommended_df["dimensi_lingkungan_sosial"] = compute_weighted_dimension(
        recommended_df,
        RECOMMENDED_INDICATOR_WEIGHTS["dimensi_lingkungan_sosial"],
    )
    recommended_df["iid_rt"] = (
        DIMENSION_WEIGHTS["akses_perangkat"] * recommended_df["dimensi_akses_perangkat"]
        + DIMENSION_WEIGHTS["konektivitas"] * recommended_df["dimensi_konektivitas"]
        + DIMENSION_WEIGHTS["kapasitas_manusia"] * recommended_df["dimensi_kapasitas_manusia"]
        + DIMENSION_WEIGHTS["penggunaan_digital"] * recommended_df["dimensi_penggunaan_digital"]
        + DIMENSION_WEIGHTS["lingkungan_sosial"] * recommended_df["dimensi_lingkungan_sosial"]
    ).clip(lower=0, upper=1)
    recommended_df["ikd_rt"] = 1 - recommended_df["iid_rt"]

    iid_cutoffs = tuple(float(recommended_df["iid_rt"].quantile(q)) for q in (0.20, 0.40, 0.60, 0.80))
    category_ranges = build_iid_category_ranges_from_cutoffs(iid_cutoffs)
    recommended_df["kategori_iid_rt"] = recommended_df["iid_rt"].apply(
        lambda value: classify_iid_rt_with_cutoffs(value, iid_cutoffs)
    )
    return recommended_df, iid_cutoffs, category_ranges


def build_recommended_variable_explanation(
    iid_cutoffs: tuple[float, float, float, float],
    category_ranges: dict[str, str],
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> pd.DataFrame:
    explanation_df = build_variable_explanation(
        school_age_min=school_age_min,
        school_age_max=school_age_max,
    ).copy()
    explanation_df["bobot_dalam_dimensi"] = pd.NA
    explanation_df["metode_agregasi"] = pd.NA

    indicator_weight_map = {
        "indikator_A": 1 / 3,
        "indikator_B": 1 / 3,
        "indikator_C": 1 / 3,
        "indikator_D": 1.00,
        "indikator_E": 0.50,
        "indikator_F": 0.50,
        "indikator_G": 0.25,
        "indikator_H": 0.25,
        "indikator_I": 0.25,
        "indikator_J": 0.25,
        "indikator_K": 1 / 3,
        "indikator_L": 1 / 3,
        "indikator_M": 1 / 3,
    }
    explanation_df["bobot_dalam_dimensi"] = explanation_df["nama_variabel"].map(indicator_weight_map)

    explanation_df.loc[explanation_df["nama_variabel"] == "dimensi_A", "aturan_skoring"] = (
        "weighted_mean(indikator_A=0,333333; indikator_B=0,333333; indikator_C=0,333333)"
    )
    explanation_df.loc[explanation_df["nama_variabel"] == "dimensi_B", "aturan_skoring"] = (
        "weighted_mean(indikator_D=1,00)"
    )
    explanation_df.loc[explanation_df["nama_variabel"] == "dimensi_C", "aturan_skoring"] = (
        "weighted_mean(indikator_E=0,50; indikator_F=0,50; jika indikator_F tidak berlaku, bobot dinormalisasi ke indikator_E)"
    )
    explanation_df.loc[explanation_df["nama_variabel"] == "dimensi_D", "aturan_skoring"] = (
        "weighted_mean(indikator_K=0,333333; indikator_L=0,333333; indikator_M=0,333333)"
    )
    explanation_df.loc[explanation_df["nama_variabel"] == "dimensi_E", "aturan_skoring"] = (
        "weighted_mean(indikator_G=0,25; indikator_H=0,25; indikator_I=0,25; indikator_J=0,25)"
    )
    explanation_df.loc[explanation_df["nama_variabel"] == "kategori_iid_rt", "aturan_skoring"] = (
        f"quintile empiris IID-RT: sangat rendah {category_ranges['sangat rendah']}; rendah {category_ranges['rendah']}; "
        f"sedang {category_ranges['sedang']}; tinggi {category_ranges['tinggi']}; sangat tinggi {category_ranges['sangat tinggi']}"
    )
    explanation_df["metode_agregasi"] = explanation_df["nama_variabel"].apply(
        lambda value: (
            "weighted_mean_indikator"
            if value in {"dimensi_A", "dimensi_B", "dimensi_C", "dimensi_D", "dimensi_E"}
            else "weighted_sum_dimensi"
            if value == "iid_rumah_tangga"
            else "quintile_empiris"
            if value == "kategori_iid_rt"
            else pd.NA
        )
    )
    explanation_df["cutoff_q1"] = pd.NA
    explanation_df["cutoff_q2"] = pd.NA
    explanation_df["cutoff_q3"] = pd.NA
    explanation_df["cutoff_q4"] = pd.NA
    category_row_mask = explanation_df["nama_variabel"].eq("kategori_iid_rt")
    explanation_df.loc[category_row_mask, "cutoff_q1"] = iid_cutoffs[0]
    explanation_df.loc[category_row_mask, "cutoff_q2"] = iid_cutoffs[1]
    explanation_df.loc[category_row_mask, "cutoff_q3"] = iid_cutoffs[2]
    explanation_df.loc[category_row_mask, "cutoff_q4"] = iid_cutoffs[3]
    return explanation_df


def build_recommended_scheme_specification(
    iid_cutoffs: tuple[float, float, float, float],
    category_ranges: dict[str, str],
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
) -> pd.DataFrame:
    explanation_df = build_recommended_variable_explanation(
        iid_cutoffs=iid_cutoffs,
        category_ranges=category_ranges,
        school_age_min=school_age_min,
        school_age_max=school_age_max,
    )
    rows = [
        {
            "level": "ringkasan",
            "komponen": "jumlah_dimensi",
            "kode": pd.NA,
            "nama": "Jumlah dimensi IID",
            "jumlah_indikator_dimensi": 5,
            "rule_data": "Akses perangkat, konektivitas internet, kapasitas manusia, penggunaan digital, lingkungan pendukung sosial",
            "skoring": pd.NA,
            "bobot_dalam_dimensi": pd.NA,
            "bobot_dimensi_ke_iid": pd.NA,
            "catatan": "Klasifikasi rekomendasi memakai quintile empiris",
        },
        {
            "level": "ringkasan",
            "komponen": "jumlah_indikator_total",
            "kode": pd.NA,
            "nama": "Jumlah indikator total",
            "jumlah_indikator_dimensi": len(INDICATOR_OUTPUT_MAP),
            "rule_data": "3 indikator akses + 1 konektivitas + 2 kapasitas + 3 penggunaan + 4 sosial",
            "skoring": pd.NA,
            "bobot_dalam_dimensi": pd.NA,
            "bobot_dimensi_ke_iid": pd.NA,
            "catatan": "Total 13 indikator",
        },
        {
            "level": "kategori",
            "komponen": "kategori_iid_rt",
            "kode": pd.NA,
            "nama": "Aturan kategori IID-RT",
            "jumlah_indikator_dimensi": pd.NA,
            "rule_data": "Kategori dibentuk dari quintile empiris skor IID-RT seluruh KK valid",
            "skoring": (
                f"sangat rendah {category_ranges['sangat rendah']}; rendah {category_ranges['rendah']}; sedang {category_ranges['sedang']}; "
                f"tinggi {category_ranges['tinggi']}; sangat tinggi {category_ranges['sangat tinggi']}"
            ),
            "bobot_dalam_dimensi": pd.NA,
            "bobot_dimensi_ke_iid": pd.NA,
            "catatan": (
                f"Q1={format_numeric_label(iid_cutoffs[0])}; Q2={format_numeric_label(iid_cutoffs[1])}; "
                f"Q3={format_numeric_label(iid_cutoffs[2])}; Q4={format_numeric_label(iid_cutoffs[3])}"
            ),
        },
    ]
    dimension_weight_by_name = {
        "Akses perangkat": DIMENSION_WEIGHTS["akses_perangkat"],
        "Konektivitas internet": DIMENSION_WEIGHTS["konektivitas"],
        "Kapasitas manusia": DIMENSION_WEIGHTS["kapasitas_manusia"],
        "Penggunaan digital": DIMENSION_WEIGHTS["penggunaan_digital"],
        "Lingkungan pendukung sosial": DIMENSION_WEIGHTS["lingkungan_sosial"],
    }
    indicator_rows = explanation_df.loc[explanation_df["nama_variabel"].str.startswith("indikator_")].copy()
    dimension_indicator_counts = indicator_rows.groupby("dimensi")["nama_variabel"].count().to_dict()
    for row in indicator_rows.itertuples(index=False):
        rows.append(
            {
                "level": "indikator",
                "komponen": row.dimensi,
                "kode": row.nama_variabel,
                "nama": row.label_konsep,
                "jumlah_indikator_dimensi": dimension_indicator_counts.get(row.dimensi, pd.NA),
                "rule_data": row.sumber_nilai,
                "skoring": row.aturan_skoring,
                "bobot_dalam_dimensi": row.bobot_dalam_dimensi,
                "bobot_dimensi_ke_iid": dimension_weight_by_name.get(row.dimensi, pd.NA),
                "catatan": "Agregasi dimensi memakai weighted mean pada indikator yang tersedia",
            }
        )

    return pd.DataFrame(rows)


def build_iid_cutoff_summary(
    valid_households: pd.DataFrame,
    iid_cutoffs: tuple[float, float, float, float],
    category_ranges: dict[str, str],
) -> pd.DataFrame:
    distribution = (
        valid_households["kategori_iid_rt"]
        .value_counts()
        .reindex(IID_RT_CATEGORY_ORDER, fill_value=0)
        .rename("jumlah_kk")
        .reset_index()
        .rename(columns={"index": "kategori_iid_rt"})
    )
    total_kk = int(distribution["jumlah_kk"].sum())
    cutoff_map = {
        "sangat rendah": (pd.NA, iid_cutoffs[0]),
        "rendah": (iid_cutoffs[0], iid_cutoffs[1]),
        "sedang": (iid_cutoffs[1], iid_cutoffs[2]),
        "tinggi": (iid_cutoffs[2], iid_cutoffs[3]),
        "sangat tinggi": (iid_cutoffs[3], pd.NA),
    }
    distribution["rentang_iid_rt"] = distribution["kategori_iid_rt"].map(category_ranges)
    distribution["batas_bawah"] = distribution["kategori_iid_rt"].map(lambda label: cutoff_map[label][0])
    distribution["batas_atas"] = distribution["kategori_iid_rt"].map(lambda label: cutoff_map[label][1])
    distribution["proporsi_kk"] = np.where(total_kk > 0, distribution["jumlah_kk"] / total_kk, 0.0)
    distribution["total_kk"] = total_kk
    distribution["metode_kategori"] = "quintile_empiris_seluruh_kk_valid"
    return distribution[
        [
            "kategori_iid_rt",
            "rentang_iid_rt",
            "batas_bawah",
            "batas_atas",
            "jumlah_kk",
            "proporsi_kk",
            "total_kk",
            "metode_kategori",
        ]
    ]


def build_scheme_comparison_summary(
    baseline_valid: pd.DataFrame,
    recommended_valid: pd.DataFrame,
    iid_cutoffs: tuple[float, float, float, float],
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    baseline_stats = baseline_valid["iid_rt"].describe(percentiles=[0.20, 0.40, 0.60, 0.80, 0.90, 0.95, 0.99])
    recommended_stats = recommended_valid["iid_rt"].describe(percentiles=[0.20, 0.40, 0.60, 0.80, 0.90, 0.95, 0.99])
    stat_map = {
        "mean": "rata_rata",
        "std": "standar_deviasi",
        "min": "minimum",
        "20%": "quintile_1",
        "40%": "quintile_2",
        "60%": "quintile_3",
        "80%": "quintile_4",
        "90%": "persentil_90",
        "95%": "persentil_95",
        "99%": "persentil_99",
        "max": "maksimum",
    }
    for stat_key, stat_label in stat_map.items():
        rows.append(
            {
                "kelompok": "statistik_iid_rt",
                "metrik": stat_label,
                "skema_awal": float(baseline_stats[stat_key]),
                "skema_rekomendasi": float(recommended_stats[stat_key]),
                "catatan": pd.NA,
            }
        )

    baseline_counts = (
        baseline_valid["iid_rt"].apply(classify_iid_rt).value_counts().reindex(IID_RT_CATEGORY_ORDER, fill_value=0)
    )
    recommended_counts = (
        recommended_valid["kategori_iid_rt"].value_counts().reindex(IID_RT_CATEGORY_ORDER, fill_value=0)
    )
    for category in IID_RT_CATEGORY_ORDER:
        rows.append(
            {
                "kelompok": "jumlah_kk_per_kategori",
                "metrik": category,
                "skema_awal": int(baseline_counts[category]),
                "skema_rekomendasi": int(recommended_counts[category]),
                "catatan": (
                    "Skema awal memakai cutoff tetap 0,20/0,40/0,60/0,80; "
                    f"skema rekomendasi memakai quintile {format_numeric_label(iid_cutoffs[0])}/"
                    f"{format_numeric_label(iid_cutoffs[1])}/{format_numeric_label(iid_cutoffs[2])}/"
                    f"{format_numeric_label(iid_cutoffs[3])}"
                ),
            }
        )
    return pd.DataFrame(rows)


def apply_excel_number_formats(workbook: object, sheets: dict[str, pd.DataFrame]) -> None:
    decimal_prefixes = (
        "indikator_",
        "dimensi_",
        "iid_",
        "ikd_",
        "gini_",
        "proporsi_",
        "persentase_",
        "rasio_",
        "r2_",
        "shapley_",
    )
    integer_prefixes = ("jumlah_",)
    decimal_exact_names = {
        "bobot_dimensi",
        "bobot_dalam_dimensi",
        "batas_bawah",
        "batas_atas",
        "nilai",
        "skema_awal",
        "skema_rekomendasi",
        "cutoff_q1",
        "cutoff_q2",
        "cutoff_q3",
        "cutoff_q4",
        "r2_iid_desa",
        "rata_rata_delta_ikd",
        "rata_rata_delta_iid",
        "rata_rata_delta_deprivasi_digital",
        "rata_rata_delta_gini_prediksi",
        "rata_rata_kenaikan_iid",
        "rata_rata_kenaikan_iid_desa",
        "rata_rata_penurunan_ikd",
        "rata_rata_penurunan_deprivasi_digital",
        "r2_skor_ikd",
        "r2_gini_desa",
        "r2_dimensi",
        "r2_total_dimensi",
        "r2_total_iid_desa",
        "r2_total_gini_desa",
    }

    for sheet_name, df in sheets.items():
        worksheet = workbook[sheet_name[:31]]
        for column_index, column_name in enumerate(df.columns, start=1):
            normalized_name = normalize_column_name(column_name)
            number_format = None
            if normalized_name.startswith(decimal_prefixes) or normalized_name in decimal_exact_names:
                number_format = EXCEL_FLOAT_FORMAT
            elif normalized_name.startswith(integer_prefixes):
                number_format = "0"

            if number_format is None:
                continue

            for row_index in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_index, column=column_index)
                if cell.value is None or cell.value == "":
                    continue
                if isinstance(cell.value, bool) or not isinstance(cell.value, Number):
                    continue
                numeric_value = float(cell.value)
                if np.isfinite(numeric_value) and np.isclose(numeric_value, round(numeric_value)):
                    cell.value = int(round(numeric_value))
                    cell.number_format = "0"
                else:
                    cell.number_format = number_format


def save_outputs(
    keluarga_output: pd.DataFrame,
    desa_summary: pd.DataFrame,
    iid_rt_distribution_desa: pd.DataFrame,
    person_distribution_df: pd.DataFrame,
    tinggi_profile_tables: dict[str, pd.DataFrame],
    excluded_households: pd.DataFrame,
    variable_explanation: pd.DataFrame,
    output_dir: Path,
    iid_category_ranges: dict[str, str] = IID_RT_CATEGORY_RANGES,
    additional_sheets: dict[str, pd.DataFrame] | None = None,
    additional_csvs: dict[str, pd.DataFrame] | None = None,
) -> tuple[dict[str, Path], list[str]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    saved_paths: dict[str, Path] = {}
    warnings: list[str] = []
    gini_distribution_summary = build_gini_distribution_summary(desa_summary)
    extra_sheets = additional_sheets or {}
    extra_csvs = additional_csvs or {}

    def write_csv_resilient(df: pd.DataFrame, filename: str, label: str) -> None:
        target_path = output_dir / filename
        prepared_df = format_dataframe_for_csv(round_numeric_dataframe(df))
        try:
            prepared_df.to_csv(target_path, index=False, encoding="utf-8-sig")
            saved_paths[label] = target_path
            return
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_path = output_dir / f"{target_path.stem}_{timestamp}{target_path.suffix}"
            prepared_df.to_csv(fallback_path, index=False, encoding="utf-8-sig")
            saved_paths[label] = fallback_path
            warnings.append(
                f"File {target_path.name} sedang terkunci atau dipakai proses lain, hasil disimpan ke {fallback_path.name}."
            )

    def write_excel_resilient(sheets: dict[str, pd.DataFrame], filename: str, label: str) -> None:
        target_path = output_dir / filename

        def _write_excel(path: Path) -> None:
            prepared_sheets = {sheet_name: round_numeric_dataframe(df) for sheet_name, df in sheets.items()}
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                for sheet_name, df in prepared_sheets.items():
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                apply_excel_number_formats(writer.book, prepared_sheets)
                add_gini_visualization_to_indeks_desa(writer.book, gini_distribution_summary)
                add_iid_rt_visualization_sheet(
                    writer.book,
                    round_numeric_dataframe(person_distribution_df),
                    round_numeric_dataframe(iid_rt_distribution_desa),
                )
                add_tinggi_profile_sheet(
                    writer.book,
                    {key: round_numeric_dataframe(df) for key, df in tinggi_profile_tables.items()},
                )

        try:
            _write_excel(target_path)
            saved_paths[label] = target_path
            return
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fallback_path = output_dir / f"{target_path.stem}_{timestamp}{target_path.suffix}"
            _write_excel(fallback_path)
            saved_paths[label] = fallback_path
            warnings.append(
                f"File {target_path.name} sedang terkunci atau dipakai proses lain, hasil disimpan ke {fallback_path.name}."
            )

    excluded_export = excluded_households.copy()
    write_csv_resilient(keluarga_output, "data_keluarga.csv", "data_keluarga")
    write_csv_resilient(desa_summary, "indeks_desa.csv", "indeks_desa")
    write_csv_resilient(variable_explanation, "penjelasan_variabel.csv", "penjelasan_variabel")
    write_csv_resilient(excluded_export, "rumah_tangga_dikeluarkan.csv", "rumah_tangga_dikeluarkan")
    write_csv_resilient(gini_distribution_summary, "sebaran_gini_desa.csv", "sebaran_gini_desa")
    for label, df in extra_csvs.items():
        write_csv_resilient(df, f"{label}.csv", label)

    workbook_sheets = {
        "data_keluarga": keluarga_output,
        "indeks_desa": desa_summary,
        "sebaran_gini_desa": gini_distribution_summary,
        "sebaran_warga_iid_rt": person_distribution_df,
        "sebaran_iid_rt_desa": iid_rt_distribution_desa,
        "penjelasan": variable_explanation,
        "kk_dikeluarkan": excluded_export,
    }
    workbook_sheets.update(extra_sheets)
    write_excel_resilient(workbook_sheets, "hasil_olahdata.xlsx", "workbook_excel")
    return saved_paths, warnings


def run_pipeline(
    input_path: Path,
    output_dir: Path,
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
    missing_threshold: float = MISSING_THRESHOLD,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Path], list[str]]:
    person_df = load_source_data(input_path)
    valid_households, excluded_households, processing_summary = build_household_index(
        person_df=person_df,
        school_age_min=school_age_min,
        school_age_max=school_age_max,
        missing_threshold=missing_threshold,
    )
    household_master = build_household_master(valid_households, excluded_households)
    keluarga_output = build_keluarga_output(
        person_df,
        household_master,
        school_age_min=school_age_min,
        school_age_max=school_age_max,
    )
    desa_summary = build_desa_summary(household_master)
    iid_rt_distribution_desa = build_iid_rt_distribution_by_desa(household_master)
    person_distribution_df = build_iid_rt_person_distribution(person_df, household_master)
    tinggi_profile_tables = build_tinggi_profile_tables(household_master)
    gini_assessment_summary, gini_contributor_detail = build_gini_assessment_tables(household_master)
    variable_explanation = build_variable_explanation(
        school_age_min=school_age_min,
        school_age_max=school_age_max,
    )
    advanced_analysis_tables = build_advanced_analysis_tables(desa_summary, variable_explanation)
    additional_tables = {
        "ringkasan_pengolahan": processing_summary,
        "sebaran_warga_iid_rt": person_distribution_df,
        "sebaran_iid_rt_desa": iid_rt_distribution_desa,
        "ringkasan_ketimpangan": gini_assessment_summary,
        "kontributor_ketimpangan": gini_contributor_detail,
    }
    additional_tables.update(advanced_analysis_tables)
    saved_paths, save_warnings = save_outputs(
        keluarga_output,
        desa_summary,
        iid_rt_distribution_desa,
        person_distribution_df,
        tinggi_profile_tables,
        excluded_households,
        variable_explanation,
        output_dir,
        additional_sheets=additional_tables,
        additional_csvs=additional_tables,
    )
    return keluarga_output, excluded_households, desa_summary, processing_summary, saved_paths, save_warnings


def run_pipeline_recommended(
    input_path: Path,
    output_dir: Path,
    school_age_min: int = SCHOOL_AGE_MIN,
    school_age_max: int = SCHOOL_AGE_MAX,
    missing_threshold: float = MISSING_THRESHOLD,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, Path], list[str]]:
    person_df = load_source_data(input_path)
    baseline_valid_households, excluded_households, processing_summary = build_household_index(
        person_df=person_df,
        school_age_min=school_age_min,
        school_age_max=school_age_max,
        missing_threshold=missing_threshold,
    )
    recommended_valid_households, iid_cutoffs, iid_category_ranges = apply_recommended_scheme(baseline_valid_households)
    household_master = build_household_master(
        recommended_valid_households,
        excluded_households,
        iid_classifier=lambda value: classify_iid_rt_with_cutoffs(value, iid_cutoffs),
    )
    keluarga_output = build_keluarga_output(
        person_df,
        household_master,
        school_age_min=school_age_min,
        school_age_max=school_age_max,
    )
    desa_summary = build_desa_summary(household_master)
    iid_rt_distribution_desa = build_iid_rt_distribution_by_desa(household_master, iid_category_ranges)
    person_distribution_df = build_iid_rt_person_distribution(person_df, household_master, iid_category_ranges)
    tinggi_profile_tables = build_tinggi_profile_tables(household_master)
    gini_assessment_summary, gini_contributor_detail = build_gini_assessment_tables(household_master)
    variable_explanation = build_recommended_variable_explanation(
        iid_cutoffs=iid_cutoffs,
        category_ranges=iid_category_ranges,
        school_age_min=school_age_min,
        school_age_max=school_age_max,
    )
    advanced_analysis_tables = build_advanced_analysis_tables(desa_summary, variable_explanation)
    scheme_specification = build_recommended_scheme_specification(
        iid_cutoffs=iid_cutoffs,
        category_ranges=iid_category_ranges,
        school_age_min=school_age_min,
        school_age_max=school_age_max,
    )
    cutoff_summary = build_iid_cutoff_summary(recommended_valid_households, iid_cutoffs, iid_category_ranges)
    comparison_summary = build_scheme_comparison_summary(
        baseline_valid=baseline_valid_households,
        recommended_valid=recommended_valid_households,
        iid_cutoffs=iid_cutoffs,
    )
    additional_tables = {
        "skema_rekomendasi": scheme_specification,
        "batas_kategori_iid_rt": cutoff_summary,
        "perbandingan_skema": comparison_summary,
        "ringkasan_pengolahan": processing_summary,
        "sebaran_warga_iid_rt": person_distribution_df,
        "sebaran_iid_rt_desa": iid_rt_distribution_desa,
        "ringkasan_ketimpangan": gini_assessment_summary,
        "kontributor_ketimpangan": gini_contributor_detail,
    }
    additional_tables.update(advanced_analysis_tables)
    saved_paths, save_warnings = save_outputs(
        keluarga_output,
        desa_summary,
        iid_rt_distribution_desa,
        person_distribution_df,
        tinggi_profile_tables,
        excluded_households,
        variable_explanation,
        output_dir,
        iid_category_ranges=iid_category_ranges,
        additional_sheets=additional_tables,
        additional_csvs=additional_tables,
    )
    return keluarga_output, excluded_households, desa_summary, processing_summary, saved_paths, save_warnings


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Hitung Indeks Inklusi Digital rumah tangga dan dusun dari data DDP.")
    parser.add_argument(
        "input_path",
        nargs="?",
        default="data_asli.csv",
        help="Path file input CSV/XLSX. Default: data_asli.csv",
    )
    parser.add_argument(
        "--output-dir",
        default="hasil_indeks_digital",
        help="Folder output untuk file CSV hasil pengolahan.",
    )
    parser.add_argument(
        "--scheme",
        choices=["baseline", "rekomendasi"],
        default="baseline",
        help="Skema perhitungan yang digunakan. `baseline` memakai skema aktif; `rekomendasi` memakai skema rumusan Codex.",
    )
    parser.add_argument("--school-age-min", type=int, default=SCHOOL_AGE_MIN, help="Batas bawah usia sekolah.")
    parser.add_argument("--school-age-max", type=int, default=SCHOOL_AGE_MAX, help="Batas atas usia sekolah.")
    parser.add_argument(
        "--missing-threshold",
        type=float,
        default=MISSING_THRESHOLD,
        help="Ambang maksimum proporsi indikator inti yang hilang per rumah tangga.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_path)
    output_dir = Path(args.output_dir)
    if args.scheme == "rekomendasi" and args.output_dir == "hasil_indeks_digital":
        output_dir = Path("hasil_indeks_digital_skema_rekomendasi_codex")

    if args.scheme == "rekomendasi":
        keluarga_output, excluded_households, desa_summary, processing_summary, saved_paths, save_warnings = run_pipeline_recommended(
            input_path=input_path,
            output_dir=output_dir,
            school_age_min=args.school_age_min,
            school_age_max=args.school_age_max,
            missing_threshold=args.missing_threshold,
        )
    else:
        keluarga_output, excluded_households, desa_summary, processing_summary, saved_paths, save_warnings = run_pipeline(
            input_path=input_path,
            output_dir=output_dir,
            school_age_min=args.school_age_min,
            school_age_max=args.school_age_max,
            missing_threshold=args.missing_threshold,
        )
    valid_count = keluarga_output["family_id"].nunique() if "family_id" in keluarga_output.columns else len(keluarga_output)

    print(f"Input      : {input_path.resolve()}")
    print(f"Output     : {output_dir.resolve()}")
    print(f"Skema      : {args.scheme}")
    print(f"RT valid   : {valid_count:,}")
    print(f"RT keluar  : {len(excluded_households):,}")
    print(f"Desa       : {desa_summary.shape[0]:,}")
    print()
    if save_warnings:
        print("Catatan penyimpanan:")
        for warning in save_warnings:
            print(f"- {warning}")
        print()
    print("File output:")
    for label, path in saved_paths.items():
        print(f"- {label}: {path.resolve()}")
    print()
    print("Ringkasan pengolahan:")
    print(processing_summary.to_string(index=False))
    print()
    print("Contoh data keluarga:")
    print(keluarga_output.head(10).to_string(index=False))
    print()
    print("Contoh hasil desa:")
    print(desa_summary.head(10).to_string(index=False))


if __name__ == "__main__":
    main()
